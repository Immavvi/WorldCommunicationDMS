import hashlib
import re
import shutil
import subprocess
import tempfile
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class ExtractedLine:
    description: str | None = None
    unit_text: str | None = None
    quantity: Decimal | None = None
    rate: Decimal | None = None
    amount: Decimal | None = None
    hsn_text: str | None = None
    oem_make: str | None = None
    model_number: str | None = None
    tax_text: str | None = None
    remarks: str | None = None
    source_page: int | None = None
    source_serial: str | None = None
    source_raw_text: str | None = None
    extraction_outcome: str = "EXTRACTED"
    extraction_issue: str | None = None
    candidate_key: str | None = None
    source_order: int | None = None
    schedule_key: str | None = None
    group_key: str | None = None
    description_raw: str | None = None
    description_normalized: str | None = None
    uom_raw: str | None = None
    uom_normalized: str | None = None
    source_page_start: int | None = None
    source_page_end: int | None = None
    extraction_method: str | None = None
    extraction_confidence: Decimal | None = None
    extraction_issues: list[str] = field(default_factory=list)


@dataclass
class ExtractedGroup:
    source_key: str
    title_raw: str
    title_normalized: str
    source_kind: str
    sequence: int
    source_page_start: int | None = None
    source_page_end: int | None = None
    source_total: Decimal | None = None
    extracted_total: Decimal | None = None
    difference: Decimal | None = None
    reconciliation_status: str = "NEEDS_REVIEW"


@dataclass
class ExtractedSchedule:
    source_key: str
    title_raw: str
    title_normalized: str
    sequence: int
    source_page_start: int | None = None
    source_page_end: int | None = None
    source_total: Decimal | None = None
    extracted_total: Decimal | None = None
    difference: Decimal | None = None
    reconciliation_status: str = "NEEDS_REVIEW"
    groups: list[ExtractedGroup] = field(default_factory=list)


@dataclass
class ExtractedLoa:
    method: str
    text: str
    loa_number: str | None = None
    tender_reference: str | None = None
    loa_date: datetime | None = None
    completion_period: str | None = None
    completion_date: datetime | None = None
    work_description: str | None = None
    contract_value: Decimal | None = None
    division_text: str | None = None
    zone_text: str | None = None
    authority_text: str | None = None
    authority_candidates: list[dict] = field(default_factory=list)
    loa_date_provenance: str | None = None
    loa_date_source: str | None = None
    completion_date_provenance: str | None = None
    boq_reconciliation: dict = field(default_factory=dict)
    lines: list[ExtractedLine] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    schedules: list[ExtractedSchedule] = field(default_factory=list)


def _clean(value: str | None) -> str | None:
    return re.sub(r"\s+", " ", value).strip(" :-\t") if value else None


def _decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    cleaned = re.sub(r"[^0-9.()-]", "", str(value)).replace("(", "-").replace(")", "")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _date(value: str | None) -> datetime | None:
    if not value:
        return None
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(value.strip(), pattern)
        except ValueError:
            continue
    return None


def _authority_designation(value: str | None) -> str | None:
    cleaned = _clean(value)
    if not cleaned:
        return None
    cleaned = re.split(
        r"(?i)(?:\bwithin\b|\bafter\b|\bfrom\s+the\s+date\b|\bdigitally\s+signed\b|"
        r"\bview\s+signature\b|\s+[.;](?:\s|$))",
        cleaned,
        maxsplit=1,
    )[0].strip(" ,:-")
    return cleaned if cleaned and len(cleaned) <= 100 else None


def _letter_header(loa_section: str) -> tuple[str | None, datetime | None]:
    source_lines = loa_section.splitlines()
    start = next(
        (
            index
            for index, line in enumerate(source_lines)
            if re.search(r"Letter\s+No\s*:", line, re.I)
        ),
        None,
    )
    if start is None:
        return None, None
    header_lines = source_lines[start : start + 8]
    number_parts: list[str] = []
    for index, line in enumerate(header_lines):
        if index and re.search(
            r"^\s*(?:M/s\b|(?:Sub|Ref|Name\s+of\s+Work|Contract\s+Value)\s*:|The\s+)",
            line,
            re.I,
        ):
            break
        left = line[:88]
        left = re.sub(r"^.*?Letter\s+No\s*:\s*", "", left, flags=re.I)
        left = re.sub(r"\s+Dated\s*:.*$", "", left, flags=re.I)
        left = re.sub(
            r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b",
            lambda match: "" if _date(match.group()) else match.group(),
            left,
        )
        cleaned = re.sub(r"\s+", " ", left).strip(" :\t")
        if cleaned and not re.fullmatch(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}", cleaned):
            number_parts.append(cleaned)
    dates = (
        _date(value)
        for value in re.findall(r"\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b", "\n".join(header_lines))
    )
    number = re.sub(r"-\s+", "-", " ".join(number_parts))
    return _clean(number), next((value for value in dates if value), None)


def normalize_extraction(method: str, text: str, lines: list[ExtractedLine]) -> ExtractedLoa:
    result = ExtractedLoa(method=method, text=text, lines=lines)
    loa_section = re.split(
        r"\b(?:Awarded\s+Quantities\s+And\s+Rates|Item\s+Breakup)\b",
        text,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    result.loa_number, header_date = _letter_header(loa_section)
    if not result.loa_number:
        labelled_loa = re.search(
            r"(?:LOA\s+(?:No\.?|Number|Reference)|Letter\s+of\s+Acceptance\s+No\.?)"
            r"\s*[:\-]\s*([^\n]+)",
            loa_section,
            re.IGNORECASE,
        )
        result.loa_number = _clean(labelled_loa.group(1)) if labelled_loa else None
    tender = re.search(r"Tender\s+No\.?\s*[:.]?\s*([A-Za-z0-9_./-]+)", loa_section, re.IGNORECASE)
    result.tender_reference = _clean(tender.group(1)) if tender else None
    if not result.tender_reference:
        labelled_tender = re.search(
            r"(?:Tender|Contract|Work\s+Tender)\s*(?:Reference|Ref\.?)\s*[:\-]\s*([^\n]+)",
            loa_section,
            re.IGNORECASE,
        )
        result.tender_reference = _clean(labelled_tender.group(1)) if labelled_tender else None
    work = re.search(
        r"Tender\s+No\.?\s*[A-Za-z0-9_./-]+.*?\bfor\s+"
        r"(.+?)(?=\n\s*(?:\d+\.|Dear\s+Sir|The\s+Competent|Sub\s*:))",
        loa_section,
        re.IGNORECASE | re.DOTALL,
    )
    if not work:
        work = re.search(
            r"(?:Name\s+of\s+Work|Work\s+Description|Description\s+of\s+Work)"
            r"\s*[:\-]\s*(.+?)(?=\n\s*[A-Z][A-Za-z /]+\s*[:\-]|\n\s*\n)",
            loa_section,
            re.IGNORECASE | re.DOTALL,
        )
    result.work_description = _clean(work.group(1)) if work else None
    date_match = re.search(
        r"(?:^|\n)\s*(?:Dated|LOA\s+Issued\s+Date|LOA\s+Date|Date\s+of\s+LOA)"
        r"\s*[:\-]?\s*(?:\n\s*)?(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
        loa_section,
        re.IGNORECASE,
    )
    result.loa_date = header_date or (_date(date_match.group(1)) if date_match else None)
    if result.loa_date:
        result.loa_date_provenance = "SOURCE_EXTRACTED"
        result.loa_date_source = "LOA header / Dated" if header_date else "Semantic LOA date label"
    completion_date = re.search(
        r"(?:Completion\s+(?:Date|by))\s*[:\-]?\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})",
        loa_section,
        re.IGNORECASE,
    )
    result.completion_date = _date(completion_date.group(1)) if completion_date else None
    if result.completion_date:
        result.completion_date_provenance = "SOURCE_EXTRACTED"
    completion = re.search(
        r"entire\s+work\s+shall\s+be\s+completed\s+within\s+"
        r"(\d+\s+(?:days?|months?|years?))\s+from\s+the\s+date\s+of\s+issue\s+of\s+"
        r"(?:Letter\s+of\s+Acceptance|LOA)",
        loa_section,
        re.IGNORECASE | re.DOTALL,
    )
    if not completion:
        completion = re.search(
            r"(?:completion\s+period\s+shall\s+be|work\s+shall\s+be\s+completed\s+within)"
            r"\s+(\d+\s+(?:days?|months?|years?))",
            loa_section,
            re.IGNORECASE,
        )
    if completion:
        duration = _clean(completion.group(1)).lower()
        result.completion_period = re.sub(r"\bmonth\b", "months", duration)
        if (
            not result.completion_date
            and result.loa_date
            and "from the date of issue" in re.sub(r"\s+", " ", completion.group(0).lower())
        ):
            result.completion_date = _add_duration(result.loa_date, result.completion_period)
            result.completion_date_provenance = "DERIVED"
        elif (
            not result.completion_date
            and not result.loa_date
            and "from the date of issue" in re.sub(r"\s+", " ", completion.group(0).lower())
        ):
            result.completion_date_provenance = "WAITING_FOR_LOA_DATE"
    else:
        labelled_completion = re.search(
            r"(?:Completion\s+Period|Period\s+of\s+Completion)\s*[:\-]\s*([^\n]+)",
            loa_section,
            re.IGNORECASE,
        )
        result.completion_period = (
            _clean(labelled_completion.group(1)) if labelled_completion else None
        )
    accepted_patterns = (
        r"total\s+cost\s+of\s+the\s+work\s+at\s+the\s+accepted\s+rates\s+works\s+out\s+to\s+"
        r"(?:Rs\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
        r"(?:Accepted\s+LOA\s+Value|Accepted\s+Contract\s+Value|Contract\s+Value)"
        r"\s*[:\-]?\s*(?:Rs\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
    )
    accepted = []
    for pattern in accepted_patterns:
        accepted.extend(re.findall(pattern, loa_section, re.IGNORECASE | re.DOTALL))
    net_bid = re.findall(
        r"Net\s+Bid\s+Value\s*(?:Rs\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
        text,
        re.IGNORECASE,
    )
    accepted_values = {
        value for value in (_decimal(item) for item in accepted) if value is not None
    }
    net_values = {value for value in (_decimal(item) for item in net_bid) if value is not None}
    if len(accepted_values) == 1:
        result.contract_value = next(iter(accepted_values))
        if net_values and any(
            abs(value - result.contract_value) > Decimal("0.01") for value in net_values
        ):
            result.warnings.append(
                "Net Bid Value does not reconcile with the explicit accepted contract value."
            )
    elif len(accepted_values) > 1:
        formatted = ", ".join(str(candidate) for candidate in sorted(accepted_values))
        result.warnings.append(f"Conflicting accepted contract values were found: {formatted}.")
    elif len(net_values) == 1:
        result.contract_value = next(iter(net_values))
    elif len(net_values) > 1:
        formatted = ", ".join(str(candidate) for candidate in sorted(net_values))
        result.warnings.append(f"Conflicting Net Bid Values were found: {formatted}.")
    division = re.search(
        r"(?:^|\n)\s*([A-Za-z][A-Za-z ]{1,40})\s+DIVISION(?:\b|[-/])",
        loa_section,
        re.IGNORECASE,
    )
    zone = re.search(
        r"(?:^|\n)\s*([A-Za-z][A-Za-z ]{1,60})\s+(?:RLY|RAILWAY|RAILWAY\s+ZONE)\b",
        loa_section,
        re.IGNORECASE,
    )
    result.division_text = _clean(division.group(1)) if division else None
    result.zone_text = _clean(zone.group(1)) if zone else None
    authority = re.search(
        r"(?:Direction\s*:.*?)directed\s+to\s+([^\n]+?)(?=\s+for\s+briefing|[.;]\s*$)",
        loa_section,
        re.IGNORECASE | re.DOTALL | re.MULTILINE,
    )
    if not authority:
        authority = re.search(
            r"(?:Consignee|Accepting\s+Authority|Issuing\s+Authority)\s*[:\-]\s*([^\n]+)",
            loa_section,
            re.IGNORECASE,
        )
    result.authority_text = _authority_designation(authority.group(1)) if authority else None
    if result.authority_text:
        result.authority_candidates.append(
            {
                "text": result.authority_text,
                "role": "EXECUTION_AUTHORITY"
                if "directed" in (authority.group(0).lower())
                else "CONSIGNEE",
                "source": "Direction clause"
                if "directed" in authority.group(0).lower()
                else "Explicit label",
            }
        )
    signatory = re.search(
        r"(?m)^\s*([^\n]{1,100}?)\s*\n\s*Digitally\s+Signed\s*$",
        loa_section,
        re.IGNORECASE,
    )
    if not signatory:
        signatory = re.search(r"(?m)^\s*([A-Za-z.]{2,20}/[A-Za-z0-9()/-]{2,50})\s*$", loa_section)
    if signatory:
        signing_text = _authority_designation(signatory.group(1))
        if signing_text and all(
            normalized.get("text", "").lower() != signing_text.lower()
            for normalized in result.authority_candidates
        ):
            result.authority_candidates.append(
                {
                    "text": signing_text,
                    "role": "ISSUING_AUTHORITY",
                    "source": "Signature block",
                }
            )
    consignee = re.search(
        r"materials?\s+(?:are|is)\s+to\s+be\s+delivered\s+at\s+the\s+stores?\s+of\s+"
        r"([^.;\n]+)",
        loa_section,
        re.IGNORECASE,
    )
    if consignee:
        consignee_text = _authority_designation(consignee.group(1))
        if consignee_text and all(
            candidate.get("text", "").lower() != consignee_text.lower()
            for candidate in result.authority_candidates
        ):
            result.authority_candidates.append(
                {
                    "text": consignee_text,
                    "role": "CONSIGNEE",
                    "source": "Material delivery / store clause",
                }
            )
    for name, value in (
        ("LOA Number", result.loa_number),
        ("LOA Issued Date", result.loa_date),
        ("Work Description", result.work_description),
        ("Contract Value", result.contract_value),
    ):
        if not value:
            result.warnings.append(f"{name} is missing or could not be extracted.")
    if not lines:
        result.warnings.append("No structured BOQ rows were detected.")
    return result


def _add_duration(start: datetime, duration: str) -> datetime:
    match = re.fullmatch(r"(\d+)\s+(days?|months?|years?)", duration)
    if not match:
        return start
    count, unit = int(match.group(1)), match.group(2)
    if unit.startswith("day"):
        return start + timedelta(days=count)
    months = count * 12 if unit.startswith("year") else count
    target_month = start.month - 1 + months
    year, month = start.year + target_month // 12, target_month % 12 + 1
    return start.replace(year=year, month=month, day=min(start.day, monthrange(year, month)[1]))


def derive_completion_date(start, duration: str):
    return _add_duration(datetime.combine(start, datetime.min.time()), duration).date()


HEADER_ALIASES = {
    "source_serial": ("item", "item no", "sl no", "serial no", "sr no", "sn no"),
    "description": ("description", "item description", "description of item", "nomenclature"),
    "unit_text": ("unit", "uom"),
    "quantity": ("quantity", "qty", "approved qty"),
    "rate": ("rate", "unit rate"),
    "amount": ("amount", "value", "total"),
    "hsn_text": ("hsn", "hsn code"),
    "oem_make": ("oem", "make"),
    "model_number": ("model", "part number", "model / part number"),
    "tax_text": ("gst", "tax"),
}


def _header_map(values: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, raw in enumerate(values):
        normalized = re.sub(r"\s+", " ", raw.strip().lower().replace(".", ""))
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                result[field_name] = index
    return result


class ExcelLoaExtractor:
    def extract(self, path: Path) -> ExtractedLoa:
        workbook = load_workbook(path, data_only=True, read_only=False)
        text_parts: list[str] = []
        lines: list[ExtractedLine] = []
        for sheet in workbook.worksheets:
            rows = [
                ["" if cell.value is None else str(cell.value) for cell in row]
                for row in sheet.iter_rows()
            ]
            text_parts.extend(" | ".join(value for value in row if value).strip() for row in rows)
            mapping: dict[str, int] | None = None
            for source_row_number, row in enumerate(rows, 1):
                candidate = _header_map(row)
                if "description" in candidate and {"quantity", "rate"}.intersection(candidate):
                    mapping = candidate
                    continue
                if not mapping or not any(row):
                    continue
                description = (
                    _clean(row[mapping["description"]])
                    if mapping["description"] < len(row)
                    else None
                )
                if not description or description.lower() in HEADER_ALIASES["description"]:
                    continue

                def value(name: str):
                    index = mapping.get(name)
                    return row[index] if index is not None and index < len(row) else None

                quantity, rate, amount = (
                    _decimal(value("quantity")),
                    _decimal(value("rate")),
                    _decimal(value("amount")),
                )
                if not any((quantity, rate, amount)):
                    if _clean(value("source_serial")):
                        lines.append(
                            ExtractedLine(
                                description=description,
                                unit_text=_clean(value("unit_text")),
                                source_serial=str(source_row_number),
                                source_raw_text=_clean(" | ".join(row)),
                                extraction_outcome="NEEDS_REVIEW",
                                extraction_issue=(
                                    "Could not reliably identify mandatory BOQ numeric fields."
                                ),
                            )
                        )
                        continue
                    if lines:
                        lines[-1].description = f"{lines[-1].description}\n{description}"
                    continue
                lines.append(
                    ExtractedLine(
                        description=description,
                        unit_text=_clean(value("unit_text")),
                        quantity=quantity,
                        rate=rate,
                        amount=amount
                        if amount is not None
                        else quantity * rate
                        if quantity is not None and rate is not None
                        else None,
                        hsn_text=_clean(value("hsn_text")),
                        oem_make=_clean(value("oem_make")),
                        model_number=_clean(value("model_number")),
                        tax_text=_clean(value("tax_text")),
                        source_serial=_clean(value("source_serial")) or str(source_row_number),
                        source_raw_text=_clean(" | ".join(row)),
                    )
                )
        result = normalize_extraction("XLSX", "\n".join(filter(None, text_parts)), lines)
        result.boq_reconciliation = {
            "schedules_detected": len(workbook.worksheets),
            "schedule_sections_detected": len(workbook.worksheets),
            "schedule_sections_accounted_for": len(workbook.worksheets),
            "document_coverage_status": "COMPLETE" if lines else "NEEDS_REVIEW",
            "item_groups_detected": 0,
            "source_rows_detected": len(lines),
            "extracted_successfully": sum(line.extraction_outcome == "EXTRACTED" for line in lines),
            "needs_review": sum(line.extraction_outcome == "NEEDS_REVIEW" for line in lines),
            "unparsed_rejected": 0,
            "explicitly_ignored": 0,
            "total_discrepancies": 0,
            "complete": bool(lines)
            and all(line.extraction_outcome == "EXTRACTED" for line in lines),
        }
        return result


class LocalOcrAdapter:
    def available(self) -> bool:
        return shutil.which("pdftoppm") is not None and shutil.which("tesseract") is not None

    def extract_pages(self, path: Path, page_numbers: list[int] | None = None) -> dict[int, str]:
        if not self.available():
            raise RuntimeError(
                "Local OCR is unavailable. Install Tesseract (`brew install tesseract`) "
                "and retry extraction; the original upload has been preserved."
            )
        with tempfile.TemporaryDirectory() as directory:
            prefix = Path(directory) / "page"
            subprocess.run(
                ["pdftoppm", "-png", "-r", "300", str(path), str(prefix)],
                check=True,
                capture_output=True,
            )
            chunks: dict[int, str] = {}
            for page_number, image in enumerate(
                sorted(Path(directory).glob("page-*.png")), start=1
            ):
                if page_numbers is not None and page_number not in page_numbers:
                    continue
                result = subprocess.run(
                    ["tesseract", str(image), "stdout", "-l", "eng", "--psm", "6"],
                    check=True,
                    capture_output=True,
                    text=True,
                )
                chunks[page_number] = result.stdout
            return chunks

    def extract(self, path: Path) -> str:
        pages = self.extract_pages(path)
        return "\f".join(pages[number] for number in sorted(pages))


class PdfLoaExtractor:
    def __init__(self, ocr: LocalOcrAdapter | None = None) -> None:
        self.ocr = ocr or LocalOcrAdapter()

    def extract(self, path: Path) -> ExtractedLoa:
        self.last_schedules = []
        self.last_reconciliation = {}
        result = subprocess.run(
            ["pdftotext", "-layout", str(path), "-"], check=True, capture_output=True, text=True
        )
        native_pages = result.stdout.split("\f")
        if native_pages and not native_pages[-1].strip():
            native_pages.pop()
        native_pages = native_pages or [result.stdout]
        usable = {
            page_number
            for page_number, page_text in enumerate(native_pages, start=1)
            if self._usable_page(page_text)
        }
        missing = [
            page_number
            for page_number in range(1, len(native_pages) + 1)
            if page_number not in usable
        ]
        ocr_pages: dict[int, str] = {}
        if missing:
            if hasattr(self.ocr, "extract_pages"):
                ocr_pages = self.ocr.extract_pages(path, missing)
            else:
                legacy_text = self.ocr.extract(path)
                legacy_pages = legacy_text.split("\f")
                ocr_pages = {
                    page_number: legacy_pages[index]
                    for index, page_number in enumerate(missing)
                    if index < len(legacy_pages)
                }
        combined_pages = [
            native_pages[index - 1] if index in usable else ocr_pages.get(index, "")
            for index in range(1, len(native_pages) + 1)
        ]
        if not any(self._usable_page(page) for page in combined_pages):
            raise RuntimeError(
                "No usable text could be extracted from the PDF after local OCR. "
                "The original upload has been preserved for retry."
            )
        method = "MIXED_PDF" if usable and ocr_pages else "OCR" if ocr_pages else "NATIVE_PDF"
        text = "\f".join(combined_pages)
        parsed_lines = self._parse_table(text)
        for line in parsed_lines:
            line.extraction_method = method
        extracted = normalize_extraction(method, text, parsed_lines)
        extracted.schedules = getattr(self, "last_schedules", [])
        reconciliation = dict(getattr(self, "last_reconciliation", {}))
        reconciliation.update(
            {
                "pages_total": len(combined_pages),
                "pages_processed": sum(bool(page.strip()) for page in combined_pages),
                "native_text_pages": len(usable),
                "ocr_pages": len(ocr_pages),
            }
        )
        schedule_totals = [
            _decimal(schedule.get("source_total"))
            for schedule in extracted.schedules
            if schedule.get("source_total") is not None
        ]
        schedule_sum = (
            sum(schedule_totals, Decimal("0"))
            if extracted.schedules and len(schedule_totals) == len(extracted.schedules)
            else None
        )
        net_values = {
            value
            for value in (
                _decimal(item)
                for item in re.findall(
                    r"Net\s+Bid\s+Value\s*(?:Rs\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)",
                    text,
                    re.I,
                )
            )
            if value is not None
        }
        net_value = next(iter(net_values)) if len(net_values) == 1 else None
        document_basis = self._document_value_basis(text)
        adjusted_schedule_totals = [
            _decimal((schedule.get("value_basis") or {}).get("adjusted_total"))
            for schedule in extracted.schedules
        ]
        adjusted_schedule_sum = (
            sum(adjusted_schedule_totals, Decimal("0"))
            if extracted.schedules and all(value is not None for value in adjusted_schedule_totals)
            else None
        )
        if extracted.contract_value is None and net_value is None and schedule_sum is not None:
            extracted.contract_value = schedule_sum
            extracted.warnings.append(
                "Contract value was derived from fully represented source schedule totals."
            )
        contract_discrepancies = 0
        if extracted.contract_value is not None and net_value is not None:
            contract_discrepancies += abs(extracted.contract_value - net_value) > Decimal("0.01")
        if document_basis and schedule_sum is not None:
            contract_discrepancies += (
                abs(document_basis["base_total"] - schedule_sum) > Decimal("0.01")
            )
        if document_basis and adjusted_schedule_sum is not None:
            contract_discrepancies += (
                abs(document_basis["adjusted_total"] - adjusted_schedule_sum)
                > Decimal("0.01")
            )
        if document_basis and net_value is not None:
            expected_net = document_basis["adjusted_total"] * (
                Decimal("1") - document_basis["rebate_percentage"] / Decimal("100")
            )
            contract_discrepancies += abs(expected_net - net_value) > Decimal("0.01")
        reconciliation["contract_value_reconciliation"] = {
            "selected_value": str(extracted.contract_value)
            if extracted.contract_value is not None
            else None,
            "net_bid_value": str(net_value) if net_value is not None else None,
            "schedule_total_sum": str(schedule_sum) if schedule_sum is not None else None,
            "adjusted_schedule_total_sum": (
                str(adjusted_schedule_sum) if adjusted_schedule_sum is not None else None
            ),
            "document_value_basis": (
                {key: str(value) for key, value in document_basis.items()}
                if document_basis
                else None
            ),
            "discrepancies": contract_discrepancies,
        }
        reconciliation["total_discrepancies"] = (
            reconciliation.get("total_discrepancies", 0) + contract_discrepancies
        )
        if contract_discrepancies:
            reconciliation["complete"] = False
        if reconciliation["pages_processed"] != reconciliation["pages_total"]:
            reconciliation["document_coverage_status"] = "NEEDS_REVIEW"
            reconciliation["complete"] = False
        extracted.boq_reconciliation = reconciliation
        return extracted

    @staticmethod
    def _usable_page(text: str) -> bool:
        return len(re.sub(r"[^A-Za-z0-9]", "", text)) >= 40

    def _parse_table(self, text: str) -> list[ExtractedLine]:
        logical = self._parse_logical_candidates(text)
        if logical is not None:
            return logical
        awarded = self._parse_awarded_quantities(text)
        if awarded:
            return awarded
        item_breakup = re.split(r"\bItem\s+Breakup\b", text, maxsplit=1, flags=re.IGNORECASE)
        if len(item_breakup) < 2:
            return []
        boq_text = re.split(
            r"\n\s*(?:SOUTH\s+EASTERN\s+RAILWAY\s*\n\s*TENDER\s+DOCUMENT|"
            r"TENDER\s+DOCUMENT\s+e-Tender)",
            item_breakup[1],
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0]
        lines: list[ExtractedLine] = []
        schedule: str | None = None
        item_group: str | None = None
        blocks = re.split(r"\n\s*\n", boq_text)
        page_number = item_breakup[0].count("\f") + 1
        group_totals: list[dict[str, str]] = []
        sections: dict[tuple[str, str], dict] = {}
        for block in blocks:
            page_number += block.count("\f")
            cleaned_lines = [self._clean_boq_line(raw) for raw in block.splitlines()]
            cleaned_lines = [raw for raw in cleaned_lines if raw]
            if not cleaned_lines:
                continue
            joined = " ".join(cleaned_lines)
            schedule_match = re.search(
                r"Schedule\s+[A-Z].*?(?=\s+Item\s*-|\s+S\s*No|$)", joined, re.IGNORECASE
            )
            if schedule_match:
                schedule = _clean(schedule_match.group())
            group_match = re.search(r"Item\s*-\s*\d+\s+.+", joined, re.IGNORECASE)
            if group_match:
                item_group = _clean(group_match.group())
            if self._is_boq_header(joined) or re.match(r"^(?:Schedule|Item\s*-)", joined, re.I):
                continue
            total_match = re.search(
                r"\bTotal\b\s*(?:Rs\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)\s*$",
                joined,
                re.IGNORECASE,
            )
            if total_match and item_group:
                source_total = _decimal(total_match.group(1))
                extracted_total = sum(
                    (line.amount or Decimal("0"))
                    for line in lines
                    if item_group in (line.remarks or "")
                    and (not schedule or schedule in (line.remarks or ""))
                    and line.extraction_outcome == "EXTRACTED"
                )
                if source_total is not None:
                    difference = source_total - extracted_total
                    group_totals.append(
                        {
                            "schedule": schedule or "",
                            "item_group": item_group,
                            "source_total": str(source_total),
                            "extracted_total": str(extracted_total),
                            "difference": str(difference),
                            "status": "MATCHED" if difference == 0 else "NEEDS_REVIEW",
                        }
                    )
                    section = sections.setdefault(
                        (schedule or "Unlabelled schedule", item_group),
                        self._new_section(schedule, item_group, page_number),
                    )
                    section["source_total"] = str(source_total)
                    section["extracted_total"] = str(extracted_total)
                    section["difference"] = str(difference)
                    section["total_status"] = "MATCHED" if difference == 0 else "NEEDS_REVIEW"
                continue
            row_indices = [
                index for index, raw in enumerate(cleaned_lines) if self._parse_boq_row(raw)
            ]
            if not row_indices:
                candidates = [raw for raw in cleaned_lines if self._looks_like_candidate(raw)]
                if candidates:
                    section = sections.setdefault(
                        (schedule or "Unlabelled schedule", item_group or "Unlabelled group"),
                        self._new_section(schedule, item_group, page_number),
                    )
                    context = " | ".join(part for part in (schedule, item_group) if part)
                    for candidate in candidates:
                        serial_match = re.match(r"^\s*(\d+)", candidate)
                        unresolved = ExtractedLine(
                            description=_clean(candidate),
                            remarks=context or None,
                            source_page=page_number,
                            source_serial=serial_match.group(1) if serial_match else None,
                            source_raw_text=_clean(candidate),
                            extraction_outcome="NEEDS_REVIEW",
                            extraction_issue=(
                                "Could not reliably identify mandatory BOQ columns."
                            ),
                        )
                        lines.append(unresolved)
                        self._count_section_line(section, unresolved, page_number)
                    continue
                if lines and not re.search(r"\bTotal\b", joined, re.IGNORECASE):
                    continuation = self._description_fragment(joined)
                    if continuation:
                        lines[-1].description = _clean(
                            f"{lines[-1].description or ''} {continuation}"
                        )
                continue
            unresolved_indices = [
                index
                for index, raw in enumerate(cleaned_lines)
                if index not in row_indices and self._looks_like_candidate(raw)
            ]
            for position, row_index in enumerate(row_indices):
                parsed = self._parse_boq_row(cleaned_lines[row_index])
                if parsed is None:
                    continue
                serial, inline_description, unit, quantity, rate, amount = parsed
                start = 0 if position == 0 else row_index
                end = (
                    row_indices[position + 1]
                    if position + 1 < len(row_indices)
                    else len(cleaned_lines)
                )
                before = (
                    [
                        raw
                        for index, raw in enumerate(cleaned_lines[start:row_index], start)
                        if index not in unresolved_indices
                    ]
                    if position == 0
                    else []
                )
                after = [
                    raw
                    for index, raw in enumerate(cleaned_lines[row_index + 1 : end], row_index + 1)
                    if index not in unresolved_indices
                ]
                description_parts = [self._description_fragment(raw) for raw in before]
                description_parts.append(inline_description)
                description_parts.extend(self._description_fragment(raw) for raw in after)
                description = _clean(" ".join(part for part in description_parts if part))
                context = " | ".join(
                    part for part in (schedule, item_group, f"S No: {serial}") if part
                )
                issue = None
                if quantity * rate != amount:
                    issue = "Source amount differs from Quantity × Rate; owner review required."
                lines.append(
                    ExtractedLine(
                        description=description,
                        unit_text=unit,
                        quantity=quantity,
                        rate=rate,
                        amount=amount,
                        remarks=context or None,
                        source_page=page_number,
                        source_serial=serial,
                        source_raw_text=_clean(" ".join(cleaned_lines)),
                        extraction_outcome="NEEDS_REVIEW" if issue else "EXTRACTED",
                        extraction_issue=issue,
                    )
                )
                section = sections.setdefault(
                    (schedule or "Unlabelled schedule", item_group or "Unlabelled group"),
                    self._new_section(schedule, item_group, page_number),
                )
                self._count_section_line(section, lines[-1], page_number)
            for unresolved_index in unresolved_indices:
                raw = cleaned_lines[unresolved_index]
                serial_match = re.match(r"^\s*(\d+)", raw)
                context = " | ".join(part for part in (schedule, item_group) if part)
                unresolved_line = ExtractedLine(
                    description=_clean(raw),
                    remarks=context or None,
                    source_page=page_number,
                    source_serial=serial_match.group(1) if serial_match else None,
                    source_raw_text=_clean(raw),
                    extraction_outcome="NEEDS_REVIEW",
                    extraction_issue="Could not reliably identify mandatory BOQ columns.",
                )
                lines.append(unresolved_line)
                section = sections.setdefault(
                    (schedule or "Unlabelled schedule", item_group or "Unlabelled group"),
                    self._new_section(schedule, item_group, page_number),
                )
                self._count_section_line(section, unresolved_line, page_number)
        schedules = {line.remarks.split(" | ")[0] for line in lines if line.remarks}
        unresolved = sum(line.extraction_outcome != "EXTRACTED" for line in lines)
        detected_schedule_headings = {
            _clean(match.group())
            for match in re.finditer(r"Schedule\s+[A-Z][^\n]*", boq_text, re.IGNORECASE)
        }
        accounted_schedules = {section["schedule"] for section in sections.values()}
        document_coverage_complete = bool(detected_schedule_headings) and all(
            any(
                normalized_heading.lower() == accounted.lower() for accounted in accounted_schedules
            )
            for normalized_heading in detected_schedule_headings
        )
        section_results = list(sections.values())
        for section in section_results:
            section["reconciliation_status"] = (
                "COMPLETE"
                if section["unresolved_row_count"] == 0
                and section.get("total_status", "MATCHED") == "MATCHED"
                else "NEEDS_REVIEW"
            )
        self.last_reconciliation = {
            "schedules_detected": len(schedules),
            "schedule_sections_detected": len(detected_schedule_headings),
            "schedule_sections_accounted_for": len(
                detected_schedule_headings.intersection(accounted_schedules)
            ),
            "document_coverage_status": (
                "COMPLETE" if document_coverage_complete else "NEEDS_REVIEW"
            ),
            "item_groups_detected": len(section_results),
            "source_rows_detected": len(lines),
            "extracted_successfully": len(lines) - unresolved,
            "needs_review": unresolved,
            "unparsed_rejected": sum(
                line.extraction_outcome == "REJECTED_WITH_REASON" for line in lines
            ),
            "missing_mandatory_structure": sum(
                line.extraction_issue == "Could not reliably identify mandatory BOQ columns."
                for line in lines
            ),
            "group_totals": group_totals,
            "sections": section_results,
            "total_discrepancies": sum(total["status"] != "MATCHED" for total in group_totals),
            "complete": document_coverage_complete
            and unresolved == 0
            and all(total["status"] == "MATCHED" for total in group_totals),
        }
        return lines

    def _parse_logical_candidates(self, text: str) -> list[ExtractedLine] | None:
        pages = text.split("\f")
        source_lines: list[dict] = []
        for page_number, page in enumerate(pages, 1):
            for order, raw in enumerate(page.splitlines()):
                source_lines.append(
                    {
                        "page": page_number,
                        "order": order,
                        "raw": raw.rstrip(),
                        "text": _clean(raw),
                        "furniture": self._is_page_furniture(raw),
                    }
                )
        regions: list[dict] = []
        source_kind: str | None = None
        schedule = "Unlabelled schedule"
        group = "Awarded Quantities And Rates"
        attached_tender = False
        current: dict | None = None

        def flush() -> None:
            nonlocal current
            if current and current["lines"]:
                regions.append(current)
            current = None

        for line in source_lines:
            value = line["text"] or ""
            if source_kind and re.match(
                r"(?:Special|General)\s+Conditions\s+of\s+Contract|Terms\s+and\s+Conditions",
                value,
                re.I,
            ):
                flush()
                source_kind = None
                continue
            if re.search(r"\bAwarded\s+Quantities\s+And\s+Rates\b", value, re.I):
                flush()
                source_kind = "AWARDED"
                group = "Awarded Quantities And Rates"
                continue
            if re.fullmatch(r"Item\s+Breakup", value, re.I):
                flush()
                source_kind = "ITEM_BREAKUP"
                group = "Unlabelled group"
                continue
            if source_kind and re.search(r"\bTENDER\s+DOCUMENT\b", value, re.I):
                flush()
                attached_tender = True
            schedule_match = re.match(r"Schedule\s*[-.:]?\s*.+", value, re.I)
            if schedule_match and not source_kind:
                source_kind = "SCHEDULE_TABLE"
            if schedule_match and "total" not in value.lower():
                flush()
                schedule = _clean(schedule_match.group()) or "Unlabelled schedule"
                group = (
                    "Awarded Quantities And Rates"
                    if source_kind == "AWARDED"
                    else "Unlabelled group"
                )
                continue
            if attached_tender or not source_kind:
                continue
            group_match = re.match(r"Item\s*-\s*\S+.*", value, re.I)
            if source_kind == "ITEM_BREAKUP" and group_match:
                flush()
                group = _clean(group_match.group()) or "Unlabelled group"
                continue
            if line["furniture"] or self._is_boq_header(value):
                continue
            if current is None:
                current = {
                    "source_kind": source_kind,
                    "schedule": schedule,
                    "group": group,
                    "lines": [],
                }
            current["lines"].append(line)
        flush()
        if not regions:
            return None

        parsed_regions: list[tuple[dict, list[ExtractedLine], Decimal | None]] = []
        for region in regions:
            candidates = self._reconstruct_region(region)
            total = self._region_total(region["lines"])
            if candidates or total is not None:
                parsed_regions.append((region, candidates, total))
        if not any(candidates for _, candidates, _ in parsed_regions):
            return None

        by_schedule: dict[str, list[tuple[dict, list[ExtractedLine], Decimal | None]]] = {}
        for parsed in parsed_regions:
            key = self._semantic_key(parsed[0]["schedule"])
            by_schedule.setdefault(key, []).append(parsed)

        selected: list[ExtractedLine] = []
        schedules: list[dict] = []
        selected_source_kinds: set[str] = set()
        for schedule_order, (schedule_key, schedule_regions) in enumerate(
            by_schedule.items(), 1
        ):
            schedule_title = max(
                (region["schedule"] for region, _, _ in schedule_regions), key=len
            )
            schedule_basis = self._schedule_bid_basis(
                schedule_title, schedule_regions
            )
            awarded = [item for item in schedule_regions if item[0]["source_kind"] == "AWARDED"]
            breakup = [
                item for item in schedule_regions if item[0]["source_kind"] == "ITEM_BREAKUP"
            ]
            direct = [
                item for item in schedule_regions if item[0]["source_kind"] == "SCHEDULE_TABLE"
            ]
            awarded_lines = [line for _, lines, _ in awarded for line in lines]
            breakup_lines = [line for _, lines, _ in breakup for line in lines]
            def is_summary(line: ExtractedLine) -> bool:
                evidence = " ".join(
                    (line.description_raw or "", line.source_raw_text or "")
                ).lower()
                return "view details" in evidence

            summary_awarded = [line for line in awarded_lines if is_summary(line)]
            summary_reference_amounts = [
                amount
                for line in summary_awarded
                if (amount := self._summary_reference_amount(line)) is not None
            ]
            resolved_summary_detail = bool(summary_awarded and breakup_lines)
            if direct:
                chosen_regions = direct
            elif awarded_lines and breakup_lines and summary_awarded:
                chosen_regions = [
                    (region, [line for line in lines if not is_summary(line)], total)
                    for region, lines, total in awarded
                ] + breakup
                chosen_regions = [item for item in chosen_regions if item[1]]
            elif not awarded_lines:
                chosen_regions = breakup
            else:
                chosen_regions = awarded
            schedule_lines: list[ExtractedLine] = []
            group_results: list[dict] = []
            for group_order, (region, candidates, source_total) in enumerate(chosen_regions, 1):
                selected_source_kinds.add(region["source_kind"])
                group_key = f"{schedule_key}:{group_order}:{region['source_kind'].lower()}"
                group_source_total = (
                    None
                    if region["source_kind"] == "AWARDED" and schedule_basis is not None
                    else source_total
                )
                extracted_total = sum(
                    (line.amount or Decimal("0"))
                    for line in candidates
                    if line.extraction_outcome == "EXTRACTED"
                )
                difference = (
                    group_source_total - extracted_total
                    if group_source_total is not None
                    else None
                )
                for line in candidates:
                    line.schedule_key = schedule_key
                    line.group_key = group_key
                schedule_lines.extend(candidates)
                group_results.append(
                    {
                        "source_key": group_key,
                        "title_raw": region["group"],
                        "title_normalized": _clean(region["group"]) or "Unlabelled group",
                        "source_kind": region["source_kind"],
                        "sequence": group_order,
                        "source_page_start": min(
                            (line.source_page_start for line in candidates), default=None
                        ),
                        "source_page_end": max(
                            (line.source_page_end for line in candidates), default=None
                        ),
                        "source_total": (
                            str(group_source_total)
                            if group_source_total is not None
                            else None
                        ),
                        "extracted_total": str(extracted_total),
                        "difference": str(difference) if difference is not None else None,
                        "reconciliation_status": (
                            "COMPLETE"
                            if not any(
                                line.extraction_outcome != "EXTRACTED" for line in candidates
                            )
                            and (difference is None or abs(difference) <= Decimal("0.01"))
                            else "NEEDS_REVIEW"
                        ),
                        "contributes_to_schedule": not (
                            resolved_summary_detail
                            and region["source_kind"] == "ITEM_BREAKUP"
                        ),
                    }
                )
            if not schedule_lines:
                continue
            selected.extend(schedule_lines)
            schedule_extracted_total = sum(
                Decimal(group["extracted_total"])
                for group in group_results
                if group["contributes_to_schedule"]
            ) + sum(summary_reference_amounts, Decimal("0"))
            if schedule_basis is not None:
                schedule_source_total = schedule_basis["base_total"]
            else:
                awarded_totals = [
                    total
                    for region, _, total in awarded
                    if total is not None
                ]
                schedule_source_total = (
                    awarded_totals[-1]
                    if awarded_totals
                    else sum(Decimal(group["source_total"]) for group in group_results)
                    if group_results
                    and all(group["source_total"] is not None for group in group_results)
                    else None
                )
            schedule_difference = (
                schedule_source_total - schedule_extracted_total
                if schedule_source_total is not None
                else None
            )
            schedules.append(
                {
                    "source_key": schedule_key,
                    "title_raw": schedule_title,
                    "title_normalized": _clean(schedule_title),
                    "sequence": schedule_order,
                    "source_page_start": min(line.source_page_start for line in schedule_lines),
                    "source_page_end": max(line.source_page_end for line in schedule_lines),
                    "source_total": (
                        str(schedule_source_total) if schedule_source_total is not None else None
                    ),
                    "extracted_total": str(schedule_extracted_total),
                    "difference": (
                        str(schedule_difference) if schedule_difference is not None else None
                    ),
                    "reconciliation_status": (
                        "COMPLETE"
                        if all(
                            group["reconciliation_status"] == "COMPLETE"
                            for group in group_results
                        )
                        and (
                            schedule_difference is None
                            or abs(schedule_difference) <= Decimal("0.01")
                        )
                        else "NEEDS_REVIEW"
                    ),
                    "value_basis": (
                        {
                            key: str(value) if isinstance(value, Decimal) else value
                            for key, value in schedule_basis.items()
                        }
                        if schedule_basis is not None
                        else None
                    ),
                    "groups": group_results,
                }
            )
        for source_order, line in enumerate(selected, 1):
            line.source_order = source_order
            identity = "|".join(
                (
                    line.schedule_key or "",
                    line.group_key or "",
                    str(line.source_page_start or ""),
                    str(line.source_page_end or ""),
                    line.source_serial or "",
                    str(source_order),
                )
            )
            line.candidate_key = hashlib.sha256(identity.encode()).hexdigest()[:32]
        unresolved = sum(line.extraction_outcome != "EXTRACTED" for line in selected)
        group_discrepancies = sum(
            group["reconciliation_status"] != "COMPLETE"
            and group["source_total"] is not None
            for schedule_result in schedules
            for group in schedule_result["groups"]
        )
        schedule_discrepancies = sum(
            schedule["source_total"] is not None
            and schedule["difference"] is not None
            and abs(Decimal(schedule["difference"])) > Decimal("0.01")
            for schedule in schedules
        )
        total_discrepancies = group_discrepancies + schedule_discrepancies
        complete = bool(selected) and unresolved == 0 and total_discrepancies == 0
        self.last_schedules = schedules
        self.last_reconciliation = {
            "semantic_parser": "LOGICAL_ROW_V1",
            "authoritative_boq_section": (
                "Awarded Quantities And Rates"
                if selected_source_kinds == {"AWARDED"}
                else "Item Breakup"
                if selected_source_kinds == {"ITEM_BREAKUP"}
                else "Per schedule/group"
            ),
            "attached_tender_excluded": attached_tender,
            "schedules_detected": len(by_schedule),
            "schedule_sections_detected": len(by_schedule),
            "schedule_sections_accounted_for": len(schedules),
            "document_coverage_status": (
                "COMPLETE" if len(schedules) == len(by_schedule) else "NEEDS_REVIEW"
            ),
            "item_groups_detected": sum(len(item["groups"]) for item in schedules),
            "source_rows_detected": len(selected),
            "extracted_successfully": len(selected) - unresolved,
            "needs_review": unresolved,
            "unparsed_rejected": 0,
            "explicitly_ignored": 0,
            "total_discrepancies": total_discrepancies,
            "schedules": schedules,
            "complete": complete and len(schedules) == len(by_schedule),
        }
        return selected

    def _reconstruct_region(self, region: dict) -> list[ExtractedLine]:
        lines = region["lines"]
        if region["source_kind"] == "AWARDED":
            def interpreter(raw: str):
                parsed = self._parse_awarded_row(raw)
                if parsed is not None or not re.search(r"\bAt\s+Par\b", raw, re.I):
                    return parsed
                return self._parse_ocr_inline_row(raw)
        elif region["source_kind"] == "SCHEDULE_TABLE":
            interpreter = self._parse_ocr_inline_row
        else:
            interpreter = self._parse_boq_row_layout
        interpreted = {
            index: parsed
            for index, line in enumerate(lines)
            if (parsed := interpreter(line["raw"])) is not None
        }
        boundary_indices = sorted(
            set(interpreted)
            | {
                index
                for index, line in enumerate(lines)
                if self._looks_like_candidate(line["text"] or "")
                or (
                    re.match(r"^\s*\d+\b", line["raw"])
                    and "view details" in (line["text"] or "").lower()
                    and re.search(r"\d[\d,]*\.\d{1,2}\s*$", line["text"] or "")
                )
            }
        )
        candidates: list[ExtractedLine] = []
        for position, anchor_index in enumerate(boundary_indices):
            next_index = (
                boundary_indices[position + 1]
                if position + 1 < len(boundary_indices)
                else len(lines)
            )
            start = (
                anchor_index
                if position or region["source_kind"] == "SCHEDULE_TABLE"
                else 0
            )
            end = next_index
            span = lines[start:end]
            parsed = interpreted.get(anchor_index)
            raw_source = "\n".join(source["raw"] for source in span).strip()
            page_start = min(source["page"] for source in span)
            page_end = max(source["page"] for source in span)
            if parsed is None:
                serial_match = re.match(r"\s*(\d+)", lines[anchor_index]["raw"])
                issue = "Mandatory BOQ numeric tail could not be interpreted."
                candidates.append(
                    ExtractedLine(
                        description=_clean(raw_source),
                        description_raw=raw_source,
                        description_normalized=_clean(raw_source),
                        remarks=f"{region['schedule']} | {region['group']}",
                        source_page=page_start,
                        source_page_start=page_start,
                        source_page_end=page_end,
                        source_serial=serial_match.group(1) if serial_match else None,
                        source_raw_text=raw_source,
                        extraction_outcome="NEEDS_REVIEW",
                        extraction_issue=issue,
                        extraction_issues=[issue],
                        extraction_confidence=Decimal("0.25"),
                        source_order=(
                            lines[anchor_index]["page"] * 100000
                            + lines[anchor_index]["order"]
                        ),
                    )
                )
                continue
            serial, inline_description, uom, quantity, rate, amount = parsed
            description_parts: list[str] = []
            for span_index, source in enumerate(span, start):
                is_anchor = span_index == anchor_index
                fragment = self._logical_description_fragment(source["raw"], is_anchor)
                if fragment:
                    description_parts.append(fragment)
                if is_anchor and inline_description:
                    description_parts.append(inline_description)
            description_raw = "\n".join(description_parts).strip() or None
            description_normalized = _clean(description_raw)
            issues: list[str] = []
            confidence = Decimal("0.95")
            if not description_raw:
                issues.append("Complete contractual description could not be reconstructed.")
                confidence -= Decimal("0.30")
            if not uom:
                issues.append("Contractual UOM could not be read reliably from the source.")
                confidence -= Decimal("0.20")
            if not serial:
                issues.append("Contractual Sn. No. could not be read reliably from the source.")
                confidence -= Decimal("0.15")
            expected_amount = quantity * rate
            tolerance = max(Decimal("0.01"), abs(amount) * Decimal("0.0001"))
            source_adjustment = self._source_percentage_adjustment(
                lines[anchor_index]["raw"], expected_amount, amount, tolerance
            )
            if abs(expected_amount - amount) > tolerance and source_adjustment is None:
                issues.append("Source amount does not reconcile with quantity × source rate.")
                confidence -= Decimal("0.25")
            candidates.append(
                ExtractedLine(
                    source_order=(
                        lines[anchor_index]["page"] * 100000
                        + lines[anchor_index]["order"]
                    ),
                    description=description_normalized,
                    description_raw=description_raw,
                    description_normalized=description_normalized,
                    unit_text=uom,
                    uom_raw=uom,
                    uom_normalized=self._normalize_uom(uom),
                    quantity=quantity,
                    rate=rate,
                    amount=amount,
                    tax_text=(
                        f"{source_adjustment}% source adjustment"
                        if source_adjustment is not None
                        else None
                    ),
                    remarks=f"{region['schedule']} | {region['group']} | S No: {serial}",
                    source_page=page_start,
                    source_page_start=page_start,
                    source_page_end=page_end,
                    source_serial=serial,
                    source_raw_text=raw_source,
                    extraction_outcome="NEEDS_REVIEW" if issues else "EXTRACTED",
                    extraction_issue=" ".join(issues) or None,
                    extraction_issues=issues,
                    extraction_confidence=max(confidence, Decimal("0")),
                )
            )
        candidates.sort(key=lambda candidate: candidate.source_order or 0)
        return candidates

    @staticmethod
    def _source_percentage_adjustment(
        raw: str, expected: Decimal, amount: Decimal, tolerance: Decimal
    ) -> Decimal | None:
        for value in re.findall(r"(?<!\d)(\d{1,2}(?:\.\d{1,2})?)(?:\s*%)?", raw):
            percentage = _decimal(value)
            if percentage is None or percentage == 0:
                continue
            adjusted = expected * (Decimal("1") + percentage / Decimal("100"))
            if abs(adjusted - amount) <= tolerance:
                return percentage
        return None

    @staticmethod
    def _is_page_furniture(raw: str) -> bool:
        value = _clean(raw) or ""
        return bool(
            re.search(r"ireps\.gov\.in", value, re.I)
            or re.fullmatch(r"Page\s+\d+(?:\s+of\s+\d+)?", value, re.I)
            or re.match(r"\d{1,2}/\d{1,2}/\d{2,4},?\s+\d{1,2}:\d{2}", value)
            or re.fullmatch(
                r"(?:SOUTH\s+EASTERN|EASTERN|NORTHERN|SOUTHERN|WESTERN|CENTRAL)"
                r"\s+RAILWAY",
                value,
                re.I,
            )
        )

    @staticmethod
    def _schedule_bid_basis(title: str, regions: list[tuple]) -> dict | None:
        evidence = " ".join(
            [title]
            + [
                line["text"] or ""
                for region, _, _ in regions
                if region["source_kind"] == "AWARDED"
                for line in region["lines"]
            ]
        )
        match = re.search(
            r"([\d,]+\.\d{2})\s+(\d+(?:\.\d+)?)\s+([\d,]+\.\d{2})",
            evidence,
        )
        if not match:
            awarded = [item for item in regions if item[0]["source_kind"] == "AWARDED"]
            awarded_lines = [line for _, lines, _ in awarded for line in lines]
            summary_amounts = [
                amount
                for line in awarded_lines
                if "view details" in (line.source_raw_text or "").lower()
                and (amount := PdfLoaExtractor._summary_reference_amount(line)) is not None
            ]
            adjusted_totals = [total for _, _, total in awarded if total is not None]
            if not summary_amounts or not adjusted_totals:
                return None
            base_total = sum(summary_amounts, Decimal("0"))
            adjusted_total = adjusted_totals[-1]
            if base_total == 0:
                return None
            delta = (adjusted_total / base_total - Decimal("1")) * Decimal("100")
            return {
                "base_total": base_total,
                "percentage": abs(delta).quantize(Decimal("0.01")),
                "direction": "ABOVE" if delta >= 0 else "BELOW",
                "adjusted_total": adjusted_total,
            }
        base_total, percentage, adjusted_total = (
            _decimal(match.group(1)),
            _decimal(match.group(2)),
            _decimal(match.group(3)),
        )
        if base_total is None or percentage is None or adjusted_total is None:
            return None
        direction_match = re.search(r"\b(Above|Below)\b", evidence, re.I)
        return {
            "base_total": base_total,
            "percentage": percentage,
            "direction": direction_match.group(1).upper() if direction_match else None,
            "adjusted_total": adjusted_total,
        }

    @staticmethod
    def _summary_reference_amount(line: ExtractedLine) -> Decimal | None:
        for raw in (line.source_raw_text or "").splitlines():
            if "view details" not in raw.lower():
                continue
            values = re.findall(r"[\d,]+\.\d{1,2}", raw)
            if values:
                return _decimal(values[-1])
        return None

    @staticmethod
    def _document_value_basis(text: str) -> dict | None:
        total_match = re.search(
            r"Total\s+Value\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})",
            text,
            re.I,
        )
        if not total_match:
            return None
        base_total = _decimal(total_match.group(1))
        adjusted_total = _decimal(total_match.group(2))
        if base_total is None or adjusted_total is None:
            return None
        rebate_match = re.search(
            r"Rebate\s+on\s+Total\s+Value\s+([\d.]+)\s+\(%\)",
            text,
            re.I,
        )
        return {
            "base_total": base_total,
            "adjusted_total": adjusted_total,
            "rebate_percentage": (
                _decimal(rebate_match.group(1)) if rebate_match else Decimal("0")
            ),
        }

    @staticmethod
    def _semantic_key(value: str) -> str:
        schedule_designator = re.match(
            r"schedule\s*[-.:]?\s*([a-z]+\d*)\b", value, re.IGNORECASE
        )
        if schedule_designator:
            return f"schedule-{schedule_designator.group(1).lower()}"
        normalized = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
        return normalized[:120] or "unlabelled-schedule"

    @staticmethod
    def _normalize_uom(value: str | None) -> str | None:
        cleaned = _clean(value)
        if not cleaned:
            return None
        aliases = {
            "number": "Numbers",
            "numbers": "Numbers",
            "nos": "Numbers",
            "meter": "Metre",
            "metre": "Metre",
            "km": "Kilometre",
            "kilometer": "Kilometre",
            "kilometre": "Kilometre",
        }
        return aliases.get(cleaned.lower(), cleaned)

    @staticmethod
    def _logical_description_fragment(raw: str, is_anchor: bool) -> str | None:
        value = raw.strip()
        if not value or re.search(
            r"\b(?:Schedule Totals|Net Bid Value|Total Value)\b", value, re.I
        ):
            return None
        if is_anchor:
            return None
        if re.search(r"S\s*No\.?\s+Item", value, re.I):
            return None
        return value

    @staticmethod
    def _region_total(lines: list[dict]) -> Decimal | None:
        values = []
        for line in lines:
            match = re.search(
                r"(?:Schedule\s+Totals|(?<!Value\s)Total)\s+(?:Rs\.?\s*)?"
                r"([\d,]+(?:\.\d{1,2})?)\s*$",
                line["text"] or "",
                re.I,
            )
            if match and (value := _decimal(match.group(1))) is not None:
                values.append(value)
        return values[-1] if values else None

    @staticmethod
    def _parse_boq_row_layout(raw: str):
        parsed = PdfLoaExtractor._parse_boq_row(raw)
        if parsed is not None:
            return parsed
        parts = [part.strip() for part in re.split(r"\s{2,}|\|", raw) if part.strip()]
        if len(parts) < 5 or not re.fullmatch(r"\d+", parts[0]):
            return None
        quantity, rate, amount = (_decimal(parts[-3]), _decimal(parts[-2]), _decimal(parts[-1]))
        if quantity is None or rate is None or amount is None:
            return None
        uom = parts[-4]
        middle = parts[1:-4]
        description = ""
        if middle:
            first = middle[0].split(maxsplit=1)
            if len(first) == 2 and re.fullmatch(r"[A-Za-z0-9_./-]+", first[0]):
                _, description = first
                description = " ".join((description, *middle[1:])).strip()
            elif len(middle) > 1 and re.fullmatch(r"[A-Za-z0-9_./-]+", middle[0]):
                description = " ".join(middle[1:])
            else:
                description = " ".join(middle)
        return parts[0], description, uom, quantity, rate, amount

    @staticmethod
    def _parse_ocr_inline_row(raw: str):
        value = re.sub(r"(?<=\d)\s+\.(?=\d)", ".", raw)
        known_uoms = {
            "each",
            "job",
            "kilometre",
            "km",
            "lot",
            "meter",
            "metre",
            "mtr",
            "mtrs",
            "no",
            "nos",
            "number",
            "numbers",
            "set",
            "boxes",
            "core",
        }
        if re.search(r"At\s+Par", value, re.I):
            normalized = re.sub(r"[|{}\[\]\\]", " ", value)
            normalized = re.sub(
                r"(?<=\d)[iIl](?=(?:Each|Job|Kilometre|KM|Lot|Met(?:er|re)|Mtrs?|Nos?|Numbers?|Set|Boxes|Core)\b)",
                " ",
                normalized,
                flags=re.I,
            )
            normalized = re.sub(r"\s+", " ", normalized).strip()
            tail = re.match(
                r"^(.*?)\s+([\d,.]+)\s+At\s+Par\s+([\d,.]+)\s*$",
                normalized,
                re.I,
            )
            if tail:
                prefix, rate_text, amount_text = tail.groups()
                uom_pattern = "|".join(sorted(known_uoms, key=len, reverse=True))
                quantity_uom = re.search(
                    rf"([\d,.]+)\s+({uom_pattern})\.?\s*$", prefix, re.I
                )
                if quantity_uom:
                    quantity = _decimal(quantity_uom.group(1))
                    rate = _decimal(rate_text)
                    amount = _decimal(amount_text)
                    if quantity is not None and rate is not None and amount is not None:
                        description_part = prefix[: quantity_uom.start()].strip()
                        serial_match = re.match(r"(\d+)\.?\s+(.*)", description_part)
                        serial = serial_match.group(1) if serial_match else None
                        description = serial_match.group(2) if serial_match else description_part
                        if serial:
                            description = re.sub(
                                rf"\s+{re.escape(serial)}\s*$", "", description
                            )
                        # IREPS OCR often collapses the physical source-code column
                        # into the description. It remains in raw evidence only.
                        description = re.sub(
                            r"\s+(?=\S*\d)[A-Za-z0-9_./-]{4,}\s*$", "", description
                        )
                        return (
                            serial,
                            description,
                            quantity_uom.group(2),
                            quantity,
                            rate,
                            amount,
                        )
        value = re.sub(r"\s*\|\s*", " ", value).strip()
        match = re.match(
            r"^(?:(\d+)\.?\s+)?(.+?)\s+([\d,.]+)\s+([\d,.]+)\s+([\d,.]+)\s*$",
            value,
        )
        if not match:
            return None
        serial, prefix, quantity_text, rate_text, amount_text = match.groups()
        prefix_parts = prefix.rsplit(maxsplit=1)
        possible_uom = prefix_parts[-1].strip(".,:()[]").lower()
        if len(prefix_parts) == 2 and possible_uom in known_uoms:
            description, uom = prefix_parts[0], prefix_parts[1]
        else:
            description, uom = prefix, None
        quantity, rate, amount = (
            _decimal(quantity_text),
            _decimal(rate_text),
            _decimal(amount_text),
        )
        if quantity is None or rate is None or amount is None:
            return None
        return serial, description, uom, quantity, rate, amount

    def _parse_awarded_quantities(self, text: str) -> list[ExtractedLine]:
        section_match = re.search(
            r"\bAwarded\s+Quantities\s+And\s+Rates\b(?P<body>.*?)"
            r"(?=\bItem\s+Breakup\b|\Z)",
            text,
            re.IGNORECASE | re.DOTALL,
        )
        if not section_match:
            return []
        body = section_match.group("body")
        lines: list[ExtractedLine] = []
        schedule: str | None = None
        sections: dict[tuple[str, str], dict] = {}
        first_page = text[: section_match.start("body")].count("\f") + 1
        for page_offset, page_text in enumerate(body.split("\f")):
            page_number = first_page + page_offset
            for raw in page_text.splitlines():
                cleaned = self._clean_boq_line(raw)
                if not cleaned:
                    continue
                schedule_match = re.search(r"Schedule\s+[^\n]+", cleaned, re.IGNORECASE)
                if schedule_match:
                    schedule = _clean(schedule_match.group())
                    continue
                parsed = self._parse_awarded_row(cleaned)
                if not parsed:
                    continue
                serial, description, unit, quantity, rate, amount = parsed
                context = " | ".join(
                    part
                    for part in (
                        schedule,
                        "Awarded Quantities And Rates",
                        f"S No: {serial}",
                    )
                    if part
                )
                if quantity * rate != amount:
                    rate = amount / quantity
                issue = None
                if not description:
                    issue = "Wrapped source description requires owner review."
                line = ExtractedLine(
                    description=description or f"Contract item {serial}",
                    unit_text=unit,
                    quantity=quantity,
                    rate=rate,
                    amount=amount,
                    remarks=context,
                    source_page=page_number,
                    source_serial=serial,
                    source_raw_text=_clean(cleaned),
                    extraction_outcome="NEEDS_REVIEW" if issue else "EXTRACTED",
                    extraction_issue=issue,
                )
                lines.append(line)
                section = sections.setdefault(
                    (schedule or "Unlabelled schedule", "Awarded Quantities And Rates"),
                    self._new_section(schedule, "Awarded Quantities And Rates", page_number),
                )
                self._count_section_line(section, line, page_number)
        if not lines:
            return []
        unresolved = sum(line.extraction_outcome != "EXTRACTED" for line in lines)
        detected_schedules = {
            _clean(match.group())
            for match in re.finditer(r"Schedule\s+[^\n]+", body, re.IGNORECASE)
            if "total" not in match.group().lower()
        }
        accounted_schedules = {section["schedule"] for section in sections.values()}
        coverage_complete = bool(detected_schedules) and all(
            any(
                detected.lower() == accounted.lower()
                for accounted in accounted_schedules
            )
            for detected in detected_schedules
        )
        section_results = list(sections.values())
        for section in section_results:
            section["reconciliation_status"] = (
                "COMPLETE" if section["unresolved_row_count"] == 0 else "NEEDS_REVIEW"
            )
        self.last_reconciliation = {
            "authoritative_boq_section": "Awarded Quantities And Rates",
            "attached_tender_excluded": True,
            "schedules_detected": len(detected_schedules),
            "schedule_sections_detected": len(detected_schedules),
            "schedule_sections_accounted_for": sum(
                any(detected.lower() == accounted.lower() for accounted in accounted_schedules)
                for detected in detected_schedules
            ),
            "document_coverage_status": (
                "COMPLETE" if coverage_complete else "NEEDS_REVIEW"
            ),
            "item_groups_detected": len(section_results),
            "source_rows_detected": len(lines),
            "extracted_successfully": len(lines) - unresolved,
            "needs_review": unresolved,
            "unparsed_rejected": 0,
            "explicitly_ignored": 0,
            "missing_mandatory_structure": 0,
            "group_totals": [],
            "sections": section_results,
            "total_discrepancies": 0,
            "complete": coverage_complete and unresolved == 0,
        }
        return lines

    @staticmethod
    def _parse_awarded_row(raw: str):
        parts = [part.strip() for part in re.split(r"\s{2,}|\|", raw) if part.strip()]
        if len(parts) >= 5 and re.fullmatch(r"\d+", parts[0]):
            amount = _decimal(parts[-1])
            combined_tail = re.fullmatch(
                r"([\d,.]+)\s+(.+?)\s+([\d,]+(?:\.\d+)?)\s+"
                r"(?:At\s+Par|[\d.]+\s*%?.*)",
                parts[-2],
                re.I,
            )
            if amount is not None and combined_tail:
                quantity = _decimal(combined_tail.group(1))
                rate = _decimal(combined_tail.group(3))
                if quantity is not None and rate is not None:
                    return (
                        parts[0],
                        " ".join(parts[1:-3]),
                        _clean(combined_tail.group(2)),
                        quantity,
                        rate,
                        amount,
                    )
            rate_match = re.match(
                r"([\d,]+(?:\.\d+)?)\s+(?:At\s+Par|[\d.]+\s*%?.*)$",
                parts[-2],
                re.I,
            )
            quantity_unit = re.fullmatch(r"([\d,.]+)\s+(.+)", parts[-3])
            code_quantity_unit = re.fullmatch(
                r"(\S+)\s+([\d,.]+)\s+(.+)", parts[-3]
            )
            if amount is not None and rate_match and (quantity_unit or code_quantity_unit):
                rate = _decimal(rate_match.group(1))
                prefix = parts[1:-3]
                if code_quantity_unit:
                    code = code_quantity_unit.group(1)
                    quantity = _decimal(code_quantity_unit.group(2))
                    unit = _clean(code_quantity_unit.group(3))
                    description = " ".join(prefix)
                else:
                    quantity = _decimal(quantity_unit.group(1))
                    unit = _clean(quantity_unit.group(2))
                    code = prefix[-1] if prefix else None
                    description = " ".join(prefix[:-1])
                if rate is not None and quantity is not None and code:
                    return (
                        parts[0],
                        description,
                        unit,
                        quantity,
                        rate,
                        amount,
                    )
        recovered = re.match(
            r"^\s*(\d+)\s+(.*?)\s+(\S+)\s+([\d,.]+)\s+([A-Za-z][A-Za-z ]*?)\s+"
            r"([\d,]+(?:\.\d+)?)\s+(?:At\s+Par|[\d.]+\s*%?.*?)\s+"
            r"([\d,]+(?:\.\d+)?)\s*$",
            raw,
            re.IGNORECASE,
        )
        if recovered:
            serial, description, code, quantity_text, unit, rate_text, amount_text = (
                recovered.groups()
            )
            quantity, rate, amount = (
                _decimal(quantity_text),
                _decimal(rate_text),
                _decimal(amount_text),
            )
            if quantity is not None and rate is not None and amount is not None:
                return serial, _clean(description), _clean(unit), quantity, rate, amount
        return None

    @staticmethod
    def _new_section(schedule: str | None, item_group: str | None, page: int) -> dict:
        return {
            "schedule": schedule or "Unlabelled schedule",
            "item_group": item_group or "Unlabelled group",
            "source_page_start": page,
            "source_page_end": page,
            "source_serial_start": None,
            "source_serial_end": None,
            "candidate_row_count": 0,
            "structured_row_count": 0,
            "unresolved_row_count": 0,
            "ignored_row_count": 0,
        }

    @staticmethod
    def _count_section_line(section: dict, line: ExtractedLine, page: int) -> None:
        section["source_page_end"] = max(section["source_page_end"], page)
        section["source_serial_start"] = section["source_serial_start"] or line.source_serial
        section["source_serial_end"] = line.source_serial
        section["candidate_row_count"] += 1
        if line.extraction_outcome == "EXTRACTED":
            section["structured_row_count"] += 1
        elif line.extraction_outcome == "EXPLICITLY_IGNORED_BY_OWNER":
            section["ignored_row_count"] += 1
        else:
            section["unresolved_row_count"] += 1

    @staticmethod
    def _clean_boq_line(raw: str) -> str | None:
        if re.search(r"ireps\.gov\.in|^\s*\d{1,2}/\d{1,2}/\d{2}", raw, re.IGNORECASE):
            return None
        return raw.replace("\f", "").rstrip()

    @staticmethod
    def _is_boq_header(value: str) -> bool:
        return bool(re.search(r"S\s*No\.?\s+Item\s+No\.?\s+Description\s+of\s+Item", value, re.I))

    @staticmethod
    def _looks_like_candidate(value: str) -> bool:
        if re.search(r"\b(?:page|dated|total|tender)\b", value, re.IGNORECASE):
            return False
        has_numeric_tail = bool(
            re.search(r"\d[\d,]*\.\d{1,2}[^A-Za-z0-9]{0,12}$", value)
        )
        number_count = len(re.findall(r"\d[\d,.]*", value))
        has_serial = bool(re.match(r"^\s*\d+[.)]?\s+\S+", value))
        has_ocr_contract_tail = bool(
            re.search(r"\bAt\s+Par\b", value, re.I)
            and re.search(
                r"(?:Each|Job|Kilometre|KM|Lot|Met(?:er|re)|Mtrs?|Nos?|Numbers?|Set|Boxes|Core)\b",
                value,
                re.I,
            )
        )
        return has_numeric_tail and (
            (has_serial and number_count >= 3)
            or (has_ocr_contract_tail and number_count >= 2)
        )

    @staticmethod
    def _parse_boq_row(raw: str):
        parts = [part.strip() for part in re.split(r"\s{2,}|\|", raw) if part.strip()]
        if len(parts) >= 6 and re.fullmatch(r"\d+", parts[0]):
            quantity, rate, amount = (
                _decimal(parts[-3]),
                _decimal(parts[-2]),
                _decimal(parts[-1]),
            )
            if quantity is not None and rate is not None and amount is not None:
                serial = parts[0]
                item_parts = parts[1:-4]
                code_and_description = item_parts[0].split(maxsplit=1)
                unit = parts[-4]
                description = " ".join(code_and_description[1:] + item_parts[1:])
                return serial, description, unit, quantity, rate, amount
        recovered = re.match(
            r"^\s*(\d+)\s+(\S+)\s+(.*?)\s+(\d+(?:\.\d+)?)\s+"
            r"([\d,]+(?:\.\d+)?)\s+([\d,]+(?:\.\d+)?)\s*$",
            raw,
        )
        if not recovered:
            return None
        serial, _source_column, prefix, quantity_text, rate_text, amount_text = (
            recovered.groups()
        )
        prefix_parts = prefix.rsplit(maxsplit=1)
        if not prefix_parts:
            return None
        description = prefix_parts[0] if len(prefix_parts) == 2 else ""
        unit = prefix_parts[-1]
        quantity, rate, amount = (
            _decimal(quantity_text),
            _decimal(rate_text),
            _decimal(amount_text),
        )
        if quantity is None or rate is None or amount is None:
            return None
        return serial, description, unit, quantity, rate, amount

    @staticmethod
    def _description_fragment(raw: str) -> str | None:
        if re.search(r"\b(?:Total|Schedule Totals|Net Bid Value)\b", raw, re.IGNORECASE):
            return None
        return raw.strip() or None
