import re
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

BROWN = "5A2D0C"
IVORY = "FAF7F0"
TAN = "C8B08D"
WHITE = "FFFFFF"
BLACK = "211A15"
THIN = Side(style="thin", color=TAN)
LOGO_PATH = Path(__file__).parent / "assets" / "world_communication_logo.png"
RUPEE = "[$₹-en-IN] #,##0.00"
TERM_LIMITS = {
    "tax_invoice": 5,
    "proforma_invoice": 7,
    "quotation": 8,
    "purchase_order": 12,
    "challan": 5,
}


def render_excel(document):
    """Render authoritative data using the supplied World Communication visual language."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_name(document.kind)
    sheet.sheet_view.showGridLines = False
    max_col = 13 if document.kind == "quotation" else 11
    _set_widths(sheet, document.kind)
    row = _header(sheet, document, max_col)
    row = _identity_and_party(sheet, document, row, max_col)
    row = _address_blocks(sheet, document, row, max_col)
    header_row = row
    row = _lines(sheet, document, row, max_col)
    row = _summary(sheet, document, row, max_col)
    row = _bank(sheet, document, row, max_col)
    row = _payment_terms(sheet, document, row, max_col)
    row, annexure_terms = _terms(sheet, document, row, max_col)
    row = _signatures(sheet, row, max_col)
    if annexure_terms:
        row = _annexure(sheet, document, annexure_terms, row, max_col)
    _footer(sheet, row, max_col)

    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.print_title_rows = f"{header_row}:{header_row}"
    sheet.print_area = f"A1:{get_column_letter(max_col)}{row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins = PageMargins(
        left=0.15, right=0.15, top=0.2, bottom=0.2, header=0.05, footer=0.05
    )
    sheet.oddFooter.center.text = "This is a Computer Generated Document  |  E. & O.E."
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.oddHeader.left.text = "WORLD COMMUNICATION"
    sheet.oddHeader.center.text = f"{document.title} - {document.identifier}"
    sheet.oddHeader.right.text = _continuation_context(document)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _sheet_name(kind):
    return {
        "quotation": "Quotation",
        "purchase_order": "Purchase Order",
        "proforma_invoice": "Performa Invoice",
        "tax_invoice": "Tax Invoice",
        "challan": "Delivery Challan",
    }[kind]


def _set_widths(sheet, kind):
    widths = (
        [5, 21, 14, 10, 8, 10, 12, 10, 11, 10, 10, 11, 4]
        if kind == "quotation"
        else [7, 23, 13, 11, 11, 11, 13, 12, 12, 14, 14]
    )
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _header(sheet, document, max_col):
    sheet.sheet_view.zoomScale = 85
    sheet.merge_cells(start_row=1, start_column=1, end_row=4, end_column=3)
    sheet.cell(1, 1, "WORLD\nCOMMUNICATION")
    sheet.cell(1, 1).font = Font(name="Arial", size=10, bold=True, color=BLACK)
    sheet.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if LOGO_PATH.exists():
        image = Image(LOGO_PATH)
        image.width = 189
        image.height = 80
        sheet.add_image(image, "A1")
    organization = document.organization or {}
    company = organization.get("legal_name") or "WORLD COMMUNICATION"
    sheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=max_col)
    sheet.cell(1, 4, str(company).upper())
    sheet.cell(1, 4).font = Font(name="Arial", size=12, bold=True, color=BLACK)
    sheet.cell(1, 4).alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells(start_row=2, start_column=4, end_row=4, end_column=max_col)
    sheet.cell(2, 4, _organization_text(organization))
    sheet.cell(2, 4).font = Font(name="Arial", size=9, bold=True, color=BLACK)
    sheet.cell(2, 4).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for row, height in ((1, 30), (2, 18), (3, 16), (4, 16)):
        sheet.row_dimensions[row].height = height
        _ivory(sheet, row, 1, max_col)
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=max_col)
    sheet.cell(5, 1, document.title)
    _band(sheet.cell(5, 1), 12)
    sheet.row_dimensions[5].height = 20
    if document.status == "DRAFT":
        sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=max_col)
        sheet.cell(6, 1, "DRAFT - NOT AN OFFICIAL ISSUED DOCUMENT")
        sheet.cell(6, 1).font = Font(name="Arial", size=11, bold=True, color="9C1C1C")
        sheet.cell(6, 1).alignment = Alignment(horizontal="center")
        _ivory(sheet, 6, 1, max_col)
        return 7
    return 6


def _identity_and_party(sheet, document, row, max_col):
    half = max_col // 2
    identity = document.identity or [("Number", document.identifier)]
    party_label, party_value = document.parties[0] if document.parties else ("PARTY", "-")
    count = max(len(identity), 5)
    start = row
    for index in range(count):
        current = row + index
        _ivory(sheet, current, 1, max_col)
        if index < len(identity):
            label, value = identity[index]
            sheet.merge_cells(start_row=current, start_column=1, end_row=current, end_column=2)
            sheet.merge_cells(start_row=current, start_column=3, end_row=current, end_column=half)
            _label(sheet.cell(current, 1), label)
            _value(
                sheet.cell(current, 3),
                value,
                bold=label
                in {
                    "WORLD COMMUNICATION DOCUMENT NUMBER",
                    "DATE",
                    "PROJECT NAME / LOA NUMBER",
                    "RAILWAY DIVISION",
                },
            )
        if index == 0:
            sheet.merge_cells(
                start_row=current, start_column=half + 1, end_row=current, end_column=max_col
            )
            heading = party_label.upper()
            if not heading.endswith("DETAILS"):
                heading = f"{heading} DETAILS"
            _section_band(sheet.cell(current, half + 1), heading)
        elif index == 1:
            sheet.merge_cells(
                start_row=current,
                start_column=half + 1,
                end_row=start + count - 1,
                end_column=max_col,
            )
            _value(sheet.cell(current, half + 1), party_value, bold=True, wrap=True)
            sheet.cell(current, half + 1).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[current].height = 18 if index else 20
    _card(sheet, start, 1, start + count - 1, half)
    _card(sheet, start, half + 1, start + count - 1, max_col)
    return start + count + 1


def _address_blocks(sheet, document, row, max_col):
    parties = document.parties[1:] if len(document.parties) > 1 else []
    left = parties[0] if parties else ("BILL TO", "-")
    right = parties[1] if len(parties) > 1 else ("SHIP TO", "-")
    half = max_col // 2
    end = row + 2
    sheet.merge_cells(start_row=row, start_column=1, end_row=end, end_column=half)
    sheet.merge_cells(start_row=row, start_column=half + 1, end_row=end, end_column=max_col)
    for column, (label, value) in ((1, left), (half + 1, right)):
        cell = sheet.cell(row, column, f"{label.upper()}:\n{value}")
        _value(cell, cell.value, bold=True, wrap=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    _card(sheet, row, 1, end, half)
    _card(sheet, row, half + 1, end, max_col)
    for current in range(row, end + 1):
        sheet.row_dimensions[current].height = 20
    return end + 2


def _lines(sheet, document, row, max_col):
    if document.kind == "purchase_order":
        spans = [(1, 1), (2, 3), (4, 4), (5, 5), (6, 6), (7, 7), (8, 8), (9, 9), (10, 10), (11, 11)]
        headers = [
            "S. No.",
            "Description",
            "OEM / Make",
            "Model / Part No.",
            "HSN/SAC",
            "Qty",
            "UOM",
            "Rate (₹)",
            "Applicable Tax",
            "Amount (₹)",
        ]
    elif document.financial:
        spans = _financial_spans(max_col)
        headers = [
                "S. No.",
                "Description of Goods / Services",
                "Model / Part No.",
                "HSN/SAC",
                "Qty",
                "Rate (₹)",
                "Amount (₹)",
            ]
    else:
        spans = _challan_spans(max_col)
        headers = [
            "S. No.",
            "Description of Goods / Materials",
            "Model / Part No.",
            "HSN/SAC",
            "Qty",
            "Unit",
            "Remarks",
        ]
    for header, (start, end) in zip(headers, spans, strict=True):
        if end > start:
            sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        _band(sheet.cell(row, start), 9)
        sheet.cell(row, start, header)
    sheet.row_dimensions[row].height = 23
    row += 1
    for line in document.lines:
        if document.kind == "purchase_order":
            values = [
                line.number,
                line.description,
                line.oem or "-",
                line.model or "-",
                line.hsn or "-",
                float(line.quantity),
                line.unit,
                float(line.rate),
                line.gst or "-",
                float(line.amount),
            ]
        elif document.financial:
            values = (
            [
                line.number,
                line.description,
                line.model or "-",
                line.hsn or "-",
                float(line.quantity),
                float(line.rate),
                float(line.amount),
            ]
            )
        else:
            values = [
                line.number,
                line.description,
                line.model or "-",
                line.hsn or "-",
                float(line.quantity),
                line.unit,
                line.remarks or "",
            ]
        quantity_index = 5 if document.kind == "purchase_order" else 4
        for value_index, (value, (start, end)) in enumerate(
            zip(values, spans, strict=True)
        ):
            if end > start:
                sheet.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            cell = sheet.cell(row, start, value)
            _value(cell, value, wrap=True)
            cell.alignment = Alignment(
                horizontal="right"
                if isinstance(value, float)
                else ("center" if start == 1 else "left"),
                vertical="top",
                wrap_text=True,
            )
            if isinstance(value, float):
                cell.number_format = "#,##0.0000" if value_index == quantity_index else RUPEE
        _block(sheet, row, 1, row, max_col)
        sheet.row_dimensions[row].height = min(
            96, max(20, ((len(line.description) + 72) // 73) * 14)
        )
        row += 1
    if not document.lines:
        _block(sheet, row, 1, row, max_col)
        row += 1
    return row


def _summary(sheet, document, row, max_col):
    half = max_col // 2
    if document.financial:
        totals = document.totals or [("Grand Total", 0)]
        height = max(5, len(totals))
        start = row
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=half)
        _section_band(sheet.cell(row, 1), "SPECIAL INSTRUCTIONS / NOTES")
        notes = (
            "\n".join(value for value in (document.special_instructions, document.notes) if value)
            or "-"
        )
        sheet.merge_cells(
            start_row=row + 1, start_column=1, end_row=row + height - 1, end_column=half
        )
        _value(sheet.cell(row + 1, 1), notes, wrap=True)
        sheet.cell(row + 1, 1).alignment = Alignment(vertical="top", wrap_text=True)
        for offset, (label, value) in enumerate(totals):
            current = row + offset
            sheet.merge_cells(
                start_row=current, start_column=half + 1, end_row=current, end_column=max_col - 2
            )
            sheet.merge_cells(
                start_row=current, start_column=max_col - 1, end_row=current, end_column=max_col
            )
            _label(sheet.cell(current, half + 1), label)
            _value(sheet.cell(current, max_col - 1), float(value))
            sheet.cell(current, max_col - 1).number_format = RUPEE
            if label == "Grand Total":
                for col in range(half + 1, max_col + 1):
                    sheet.cell(current, col).fill = PatternFill("solid", fgColor=BROWN)
                    sheet.cell(current, col).font = Font(
                        name="Arial", size=9, bold=True, color=WHITE
                    )
        _block(sheet, start, 1, start + height - 1, max_col)
        row = start + height
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        sheet.cell(row, 1, f"AMOUNT IN WORDS: {document.amount_in_words or '-'}")
        _value(sheet.cell(row, 1), sheet.cell(row, 1).value, bold=True, wrap=True)
        _block(sheet, row, 1, row, max_col)
        sheet.row_dimensions[row].height = 30
        return row + 2

    start = row
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=half)
    _section_band(sheet.cell(row, 1), "SPECIAL INSTRUCTIONS / NOTES")
    sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 5, end_column=half)
    _value(
        sheet.cell(row + 1, 1), document.special_instructions or document.notes or "-", wrap=True
    )
    sheet.cell(row + 1, 1).alignment = Alignment(vertical="top", wrap_text=True)
    labels = (
        ("TOTAL QUANTITY", sum(float(line.quantity) for line in document.lines)),
        ("MATERIAL STATUS", "Non-Financial Delivery Challan"),
        ("REFERENCE", document.identifier),
    )
    for offset, (label, value) in enumerate(labels):
        current = row + offset
        sheet.merge_cells(
            start_row=current, start_column=half + 1, end_row=current, end_column=max_col - 2
        )
        sheet.merge_cells(
            start_row=current, start_column=max_col - 1, end_row=current, end_column=max_col
        )
        _label(sheet.cell(current, half + 1), label)
        _value(sheet.cell(current, max_col - 1), value, bold=True)
    sheet.merge_cells(start_row=row + 3, start_column=half + 1, end_row=row + 3, end_column=max_col)
    _section_band(sheet.cell(row + 3, half + 1), "RECEIVER'S ACKNOWLEDGEMENT")
    sheet.merge_cells(start_row=row + 4, start_column=half + 1, end_row=row + 5, end_column=max_col)
    _value(
        sheet.cell(row + 4, half + 1),
        "Received the above material in good condition.\n"
        "Name: __________________  Designation: __________________\n"
        "Signature / Stamp: __________________  Date: __________",
        wrap=True,
    )
    _block(sheet, start, 1, start + 5, max_col)
    sheet.row_dimensions[start + 4].height = 32
    return start + 7


def _bank(sheet, document, row, max_col):
    if document.kind not in {"proforma_invoice", "tax_invoice"} or not document.bank:
        return row
    end = row + 2
    bank = document.bank or {}
    bank_text = "BANK DETAILS\n" + (
        "\n".join(
            f"{key.replace('_', ' ').title()}: {value}" for key, value in bank.items() if value
        )
        or "Not recorded in this document snapshot"
    )
    sheet.merge_cells(start_row=row, start_column=1, end_row=end, end_column=max_col)
    _value(sheet.cell(row, 1), bank_text, bold=True, wrap=True)
    sheet.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
    _card(sheet, row, 1, end, max_col)
    return end + 2


def _payment_terms(sheet, document, row, max_col):
    payment = document.payment_terms or {}
    text = payment.get("description") or payment.get("name")
    if not text:
        return row
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _section_band(sheet.cell(row, 1), "PAYMENT TERMS")
    sheet.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=max_col)
    _value(sheet.cell(row + 1, 1), text, wrap=True)
    _card(sheet, row, 1, row + 1, max_col)
    return row + 3


def _signatures(sheet, row, max_col):
    half = max_col // 2
    end = row + 4
    sheet.merge_cells(start_row=row, start_column=1, end_row=end, end_column=half)
    sheet.merge_cells(start_row=row, start_column=half + 1, end_row=end, end_column=max_col)
    signature = (
        "FOR WORLD COMMUNICATION\n\n\n"
        "____________________________\nAuthorised Signatory"
    )
    _value(sheet.cell(row, 1), signature, bold=True, wrap=True)
    _value(
        sheet.cell(row, half + 1),
        signature,
        bold=True,
        wrap=True,
    )
    sheet.cell(row, 1).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    sheet.cell(row, half + 1).alignment = Alignment(
        horizontal="center", vertical="top", wrap_text=True
    )
    _card(sheet, row, 1, end, half)
    _card(sheet, row, half + 1, end, max_col)
    return end + 2


def _terms(sheet, document, row, max_col):
    terms = _term_items(document.terms)
    if not terms:
        return row, []
    limit = TERM_LIMITS.get(document.kind, 5)
    annexure = terms if len(terms) > limit else []
    visible = terms[:limit]
    if annexure:
        visible.append("Detailed Terms & Conditions: Refer Annexure-A.")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _section_band(sheet.cell(row, 1), "TERMS & CONDITIONS")
    row += 1
    for index, text in enumerate(visible, 1):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_col)
        _value(sheet.cell(row, 1), index, bold=True)
        _value(sheet.cell(row, 2), text, wrap=True)
        _block(sheet, row, 1, row, max_col)
        sheet.row_dimensions[row].height = min(120, max(18, ((len(text) + 155) // 156) * 15))
        row += 1
    return row + 1, annexure


def _annexure(sheet, document, terms, row, max_col):
    sheet.row_breaks.append(row - 1)
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _section_band(sheet.cell(row, 1), "ANNEXURE-A - DETAILED TERMS & CONDITIONS")
    row += 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _value(
        sheet.cell(row, 1),
        f"Document Type: {document.title}   |   Document Number: {document.identifier}   |   "
        f"Document Date: {_document_date(document)}",
        bold=True,
        wrap=True,
    )
    _card(sheet, row, 1, row, max_col)
    row += 1
    for index, text in enumerate(terms, 1):
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_col)
        _value(sheet.cell(row, 1), index, bold=True)
        _value(sheet.cell(row, 2), text, wrap=True)
        _block(sheet, row, 1, row, max_col)
        sheet.row_dimensions[row].height = min(120, max(18, ((len(text) + 155) // 156) * 15))
        row += 1
    return row + 1


def _footer(sheet, row, max_col):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    sheet.cell(row, 1, "This is a Computer Generated Document   |   E. & O.E.")
    _value(sheet.cell(row, 1), sheet.cell(row, 1).value, bold=True)
    sheet.cell(row, 1).alignment = Alignment(horizontal="center")
    _block(sheet, row, 1, row, max_col)


def _financial_spans(max_col):
    if max_col == 13:
        return [(1, 1), (2, 4), (5, 6), (7, 8), (9, 9), (10, 11), (12, 13)]
    return [(1, 1), (2, 4), (5, 6), (7, 7), (8, 8), (9, 9), (10, 11)]


def _challan_spans(max_col):
    return [(1, 1), (2, 4), (5, 6), (7, 7), (8, 8), (9, 9), (10, max_col)]


def _organization_text(value):
    address = value.get("address") or value.get("registered_address") or value.get("address_line_1")
    pieces = [
        address,
        value.get("email"),
        value.get("phone"),
        f"GSTIN: {value.get('gstin')}" if value.get("gstin") else None,
        value.get("website"),
    ]
    return "  |  ".join(str(piece) for piece in pieces if piece)


def _ivory(sheet, row, start, end):
    for column in range(start, end + 1):
        sheet.cell(row, column).fill = PatternFill("solid", fgColor=IVORY)


def _band(cell, size=10):
    cell.fill = PatternFill("solid", fgColor=BROWN)
    cell.font = Font(name="Arial", size=size, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _section_band(cell, text):
    cell.value = text
    _band(cell, 9)


def _label(cell, value):
    cell.value = str(value)
    cell.font = Font(name="Arial", size=8, bold=True, color=BLACK)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _value(cell, value, *, bold=False, wrap=False):
    cell.value = value
    cell.font = Font(name="Arial", size=8, bold=bold, color=BLACK)
    cell.fill = PatternFill("solid", fgColor=IVORY)
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)


def _block(sheet, start_row, start_col, end_row, end_col):
    for row in range(start_row, end_row + 1):
        for column in range(start_col, end_col + 1):
            if sheet.cell(row, column).fill.fill_type is None:
                sheet.cell(row, column).fill = PatternFill("solid", fgColor=IVORY)
            sheet.cell(row, column).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _card(sheet, start_row, start_col, end_row, end_col):
    medium = Side(style="medium", color=BROWN)
    for row in range(start_row, end_row + 1):
        for column in range(start_col, end_col + 1):
            cell = sheet.cell(row, column)
            if cell.fill.fill_type is None:
                cell.fill = PatternFill("solid", fgColor=IVORY)
            cell.border = Border(
                left=medium if column == start_col else Side(style=None),
                right=medium if column == end_col else Side(style=None),
                top=medium if row == start_row else Side(style=None),
                bottom=medium if row == end_row else THIN,
            )


def _term_items(text):
    if not text:
        return []
    normalized = str(text).replace("\r\n", "\n").strip()
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    if len(lines) > 1:
        return lines
    return [item.strip() for item in re.split(r"(?=\b\d+[.)]\s+)", normalized) if item.strip()]


def _document_date(document):
    for label, value in document.identity:
        if label == "DATE":
            return str(value)
    return "-"


def _continuation_context(document):
    for label, value in document.identity:
        if label == "PROJECT NAME / LOA NUMBER":
            text = f"Project / LOA: {value}"
            return text if len(text) <= 45 else f"{text[:42]}..."
    return ""
