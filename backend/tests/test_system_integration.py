from io import BytesIO
from uuid import UUID

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from test_invoicing import (
    action as invoice_action,
)
from test_invoicing import (
    invoice_foundation,
    invoice_payload,
)
from test_invoicing import (
    test_railway_invoice_preserves_contract_and_confidentiality as build_railway_chain,
)
from test_payments import issued_invoice
from test_receiving import auth, login

from app.models.auth import AuditLog
from app.models.dispatch import SupplyChallan
from app.models.invoicing import TaxInvoice


async def _pay_invoice(client, admin, super_admin, invoice):
    payment_response = await client.post(
        "/api/v1/payments",
        json={
            "receipt_date": "2026-08-29",
            "customer_party_id": str(invoice.customer_party_id),
            "organization_id": str(invoice.organization_id),
            "bank_account_id": str(invoice.bank_account_id),
            "payment_mode": "NEFT",
            "transaction_reference": f"UAT-{invoice.invoice_number}",
            "transaction_date": "2026-08-29",
            "amount_received": str(invoice.grand_total),
            "currency": "INR",
        },
        headers=auth(admin),
    )
    assert payment_response.status_code == 201, payment_response.text
    payment = payment_response.json()
    allocated = await client.post(
        f"/api/v1/payments/{payment['id']}/allocations",
        json={
            "tax_invoice_id": str(invoice.id),
            "allocated_amount": str(invoice.grand_total),
            "allocation_date": "2026-08-29",
        },
        headers=auth(admin),
    )
    assert allocated.status_code == 200, allocated.text
    confirmed = await client.post(
        f"/api/v1/payments/{payment['id']}/actions",
        json={"action": "CONFIRM", "reason": "Integration settlement"},
        headers=auth(super_admin),
    )
    assert confirmed.status_code == 200, confirmed.text
    return payment


@pytest.mark.asyncio
async def test_complete_railway_chain_reconciles_documents_payment_reports_and_audit(
    client: AsyncClient,
) -> None:
    # This established Railway setup traverses Master Data -> Project/LOA -> PO ->
    # GRN -> Challan -> PI -> Tax Invoice using one isolated database transaction graph.
    await build_railway_chain(client)
    admin = await login(client, "admin@example.com", "admin-user-password")
    super_admin = await login(client, "superadmin@example.com", "super-admin-password")
    async with client._session_factory() as session:
        invoice = await session.scalar(select(TaxInvoice))
        challan = await session.scalar(select(SupplyChallan))
        assert invoice is not None and challan is not None
        invoice_id = invoice.id
        project_id = invoice.project_id
        loa_id = invoice.loa_id
        invoice_total = invoice.grand_total

    assert (await invoice_action(client, admin, invoice_id, "SUBMIT")).status_code == 200
    assert (await invoice_action(client, super_admin, invoice_id, "APPROVE")).status_code == 200
    assert (await invoice_action(client, super_admin, invoice_id, "ISSUE")).status_code == 200
    async with client._session_factory() as session:
        invoice = await session.get(TaxInvoice, invoice_id)
    await _pay_invoice(client, admin, super_admin, invoice)

    receivable = await client.get(
        f"/api/v1/receivables/{invoice_id}", headers=auth(super_admin)
    )
    assert receivable.status_code == 200
    assert receivable.json()["invoice_total"] == f"{invoice_total:.2f}"
    assert receivable.json()["outstanding_amount"] == "0.00"
    assert receivable.json()["payment_status"] == "PAID"

    reconciliation = await client.get(
        f"/api/v1/reports/loas/{loa_id}/reconciliation", headers=auth(admin)
    )
    assert reconciliation.status_code == 200
    row = reconciliation.json()[0]
    assert row["current_approved_quantity"] == row["po_committed"]
    assert row["po_committed"] == row["received"]
    assert row["received"] == row["dispatched"]
    assert row["invoiced_quantity"] == 1
    assert row["remaining_billing"] == 4

    summary = await client.get(
        f"/api/v1/reports/projects/{project_id}/summary", headers=auth(super_admin)
    )
    assert summary.status_code == 200
    assert len(summary.json()["purchase_orders"]) == 1
    assert len(summary.json()["receipts"]) == 1
    assert len(summary.json()["challans"]) == 1
    assert len(summary.json()["tax_invoices"]) == 1
    assert summary.json()["receivables"][0]["outstanding"] == 0

    dashboard = await client.get("/api/v1/dashboard", headers=auth(super_admin))
    assert dashboard.status_code == 200
    assert dashboard.json()["financial"]["outstanding"] == 0
    evaluated = await client.post("/api/v1/alerts/evaluate", headers=auth(super_admin))
    assert evaluated.status_code == 200

    challan_api = await client.get(
        f"/api/v1/supply-challans/{challan.id}", headers=auth(admin)
    )
    forbidden = ("unit_rate", "purchase gst", "vendor cost", "margin", "profit")
    assert not any(term in challan_api.text.lower() for term in forbidden)
    for output_format in ("pdf", "excel"):
        exported = await client.get(
            f"/api/v1/documents/supply-challan/{challan.id}/{output_format}",
            headers=auth(admin),
        )
        assert exported.status_code == 200
        if output_format == "pdf":
            assert exported.content.startswith(b"%PDF-")
            continue
        else:
            sheet = load_workbook(BytesIO(exported.content), data_only=True).active
            text = " ".join(
                str(cell.value)
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            ).lower()
        assert not any(term in text for term in forbidden)

    async with client._session_factory() as session:
        actions = set(await session.scalars(select(AuditLog.action)))
        audit_count = await session.scalar(select(func.count()).select_from(AuditLog))
    assert {"create", "submit", "approve", "issue", "confirm"}.issubset(actions)
    assert audit_count and audit_count >= 20


@pytest.mark.asyncio
async def test_complete_non_railway_billing_and_settlement_needs_no_railway_data(
    client: AsyncClient,
) -> None:
    admin = await login(client, "admin@example.com", "admin-user-password")
    super_admin = await login(client, "superadmin@example.com", "super-admin-password")
    _, response = await issued_invoice(client, admin, super_admin, quantity="2")
    async with client._session_factory() as session:
        invoice = await session.get(TaxInvoice, UUID(response["id"]))
        assert invoice.business_scope == "NON_RAILWAY"
        assert invoice.railway_division_id is None
        assert invoice.railway_authority_id is None
    await _pay_invoice(client, admin, super_admin, invoice)
    receivable = (
        await client.get(f"/api/v1/receivables/{response['id']}", headers=auth(admin))
    ).json()
    assert receivable["payment_status"] == "PAID"
    assert receivable["outstanding_amount"] == "0.00"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("destination_state", "destination_code", "expected_mode"),
    (
        ("Bihar", "10", "INTRA_STATE"),
        ("West Bengal", "19", "INTER_STATE"),
        ("Karnataka", "29", "INTER_STATE"),
    ),
)
async def test_bihar_gst_switches_authoritatively_by_place_of_supply(
    client: AsyncClient,
    destination_state: str,
    destination_code: str,
    expected_mode: str,
) -> None:
    admin = await login(client, "admin@example.com", "admin-user-password")
    super_admin = await login(client, "superadmin@example.com", "super-admin-password")
    data = await invoice_foundation(client, admin, super_admin, expected_mode)
    gst_update = await client.patch(
        f"/api/v1/master-data/gst-registrations/{data['organization_gst']['id']}",
        json={"state": "Bihar", "state_code": "10"},
        headers=auth(super_admin),
    )
    assert gst_update.status_code == 200, gst_update.text
    address_update = await client.patch(
        f"/api/v1/master-data/party-addresses/{data['address']['id']}",
        json={"state": destination_state, "state_code": destination_code},
        headers=auth(admin),
    )
    assert address_update.status_code == 200, address_update.text

    created = await client.post(
        "/api/v1/tax-invoices",
        json=invoice_payload(data, quantity="1", tax_mode=expected_mode),
        headers=auth(admin),
    )
    assert created.status_code == 201, created.text
    invoice = created.json()
    assert invoice["tax_mode"] == expected_mode
    if expected_mode == "INTRA_STATE":
        assert invoice["cgst_amount"] == "8.10"
        assert invoice["sgst_amount"] == "8.10"
        assert invoice["igst_amount"] == "0.00"
    else:
        assert invoice["cgst_amount"] == "0.00"
        assert invoice["sgst_amount"] == "0.00"
        assert invoice["igst_amount"] == "16.20"
