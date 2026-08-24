from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import KeepTogether, PageBreak, Table
from test_receiving import auth, login

from app.document_engine.adapters import RenderDocument, RenderLine, purchase_order
from app.document_engine.excel import render_excel
from app.document_engine.formatting import currency, indian_number, safe_filename
from app.document_engine.pdf import (
    _continuation_parts,
    _presented_terms,
    _term_items,
    _terms_flowables,
    render_pdf,
)
from app.document_engine.service import DocumentExportService


def sample_document(kind, *, draft=False, financial=True, stress=False):
    description = "Industrial managed network appliance with extended environmental qualification "
    if stress:
        description *= 8
    lines = [
        RenderLine(
            number=index,
            description=f"{description} {index}",
            hsn="85176290",
            unit="Nos",
            quantity="123456.7890",
            rate="9876543.21" if financial else None,
            discount="5.0000" if financial else None,
            taxable="1158179004669.02" if financial else None,
            gst="IGST 18%" if financial else None,
            amount="1366651225509.44" if financial else None,
            remarks="Packed securely",
            model="C9200L-24P",
            oem="Cisco",
        )
        for index in range(1, 31 if stress else 3)
    ]
    return RenderDocument(
        kind=kind,
        title={
            "quotation": "QUOTATION",
            "purchase_order": "PURCHASE ORDER",
            "proforma_invoice": "PROFORMA INVOICE",
            "tax_invoice": "TAX INVOICE",
            "challan": "SUPPLY CHALLAN",
        }[kind],
        identifier={
            "quotation": "QTN-000001-R0",
            "purchase_order": "PO-000001",
            "proforma_invoice": "PI-000001",
            "tax_invoice": "INV-000001",
            "challan": "CH-000001",
        }[kind],
        status="DRAFT" if draft else "ISSUED",
        organization={
            "legal_name": "World Communication",
            "gstin": "19ABCDE1234F1Z5",
            "email": "office@example.com",
            "address": "Girls School Road, Near Ram Mandir, Katihar, Bihar 854105",
        },
        identity=_sample_identity(kind),
        parties=_sample_parties(kind),
        lines=lines,
        financial=financial,
        totals=[
            ("Taxable Amount", "1158179004669.02"),
            ("IGST", "208472220840.42"),
            ("Grand Total", "1366651225509.44"),
        ]
        if financial
        else [],
        amount_in_words="Indian Rupees Thirteen Kharab Sixty Six Arab Only" if financial else None,
        bank={"bank_name": "WCDMS Bank", "account_number": "1234567890", "ifsc": "WCDM0000001"}
        if kind in {"proforma_invoice", "tax_invoice"}
        else None,
        terms=("1. Delivery is subject to inspection.\n2. Warranty terms apply.\n" * 80)
        if stress
        else "1. Delivery is subject to inspection.",
        continuation_reference={
            "quotation": "REVISION 0",
            "purchase_order": "LOA: ER/LOA/2026/004",
            "proforma_invoice": "LOA: ER/LOA/2026/004",
            "tax_invoice": "LOA: ER/LOA/2026/004",
            "challan": "LOA: ER/LOA/2026/004",
        }[kind],
    )


def _sample_identity(kind):
    common = [
        ("WORLD COMMUNICATION DOCUMENT NUMBER", "DOC/1"),
        ("DATE", "2026-08-28"),
    ]
    additions = {
        "quotation": [("Enquiry / Tender", "RLY/TENDER/2026/17"), ("Validity", "2026-09-30")],
        "purchase_order": [
            (
                "PROJECT NAME / LOA NUMBER",
                "Eastern Railway Network Upgrade / ER/LOA/2026/004",
            ),
            ("RAILWAY DIVISION", "Howrah"),
            ("PROCUREMENT REQUIREMENT REFERENCE", "PR-000014"),
        ],
        "proforma_invoice": [
            (
                "PROJECT NAME / LOA NUMBER",
                "Eastern Railway Network Upgrade / ER/LOA/2026/004",
            ),
            ("RAILWAY DIVISION", "Howrah"),
        ],
        "tax_invoice": [
            (
                "PROJECT NAME / LOA NUMBER",
                "Eastern Railway Network Upgrade / ER/LOA/2026/004",
            ),
            ("PLACE OF SUPPLY", "West Bengal (19)"),
            ("PI / CHALLAN REFERENCES", "PI-000001 / CH-000001"),
            ("DUE DATE", "2026-09-27"),
        ],
        "challan": [
            (
                "PROJECT NAME / LOA NUMBER",
                "Eastern Railway Network Upgrade / ER/LOA/2026/004",
            ),
            ("RAILWAY DIVISION", "Howrah"),
            ("TRANSPORTER / VEHICLE", "Rail Logistics / WB01AB1234"),
        ],
    }
    return [*common, *additions[kind]]


def _sample_parties(kind):
    delivery = (
        "A very long delivery address, Railway Stores Depot, Industrial Area, "
        "Kolkata, West Bengal, India 700001"
    )
    if kind == "purchase_order":
        return [
            (
                "Vendor / Supplier",
                "XYZ Distributor Pvt Ltd\nGSTIN: 19AAAAA0000A1Z5\nPAN: AAAAA0000A",
            ),
            ("Buyer / Bill To", "World Communication\nGSTIN: 19ABCDE1234F1Z5"),
            ("Ship To", delivery),
        ]
    if kind == "challan":
        return [
            ("Customer", "Indian Railways"),
            ("Consignee", "Senior Divisional Signal Engineer, Howrah"),
            ("Delivery Address", delivery),
            ("Dispatch From", "World Communication, Katihar"),
        ]
    parties = [
        ("Customer", "Indian Railways\nGSTIN 19AAAAA0000A1Z5"),
        ("Bill To", "Indian Railways, Fairlie Place, Kolkata"),
        ("Ship To", delivery),
    ]
    if kind == "proforma_invoice":
        parties.append(("Railway Authority", "Senior Divisional Signal Engineer, Howrah"))
    return parties


@pytest.mark.parametrize(
    ("kind", "financial"),
    [
        ("quotation", True),
        ("purchase_order", True),
        ("proforma_invoice", True),
        ("tax_invoice", True),
        ("challan", False),
    ],
)
def test_pdf_and_excel_generation_for_document_family(kind, financial):
    document = sample_document(kind, draft=kind == "quotation", financial=financial)
    pdf = render_pdf(document)
    excel = render_excel(document)
    assert pdf.startswith(b"%PDF-")
    assert excel.startswith(b"PK\x03\x04")
    workbook = load_workbook(BytesIO(excel), data_only=False)
    sheet = workbook.active
    assert sheet.page_setup.orientation == "landscape"
    assert str(sheet.page_setup.paperSize) == sheet.PAPERSIZE_A4
    assert sheet.sheet_properties.pageSetUpPr.fitToPage is True
    values = " ".join(str(cell.value or "") for row in sheet.iter_rows() for cell in row)
    if kind == "quotation":
        assert "DRAFT - NOT AN OFFICIAL ISSUED DOCUMENT" in values
    if financial:
        assert "AMOUNT IN WORDS" in values and "Grand Total" in values
        assert "₹" in sheet["F10"].number_format or any(
            "₹" in cell.number_format for row in sheet.iter_rows() for cell in row
        )
    else:
        header_values = next(
            [cell.value for cell in row if cell.value]
            for row in sheet.iter_rows()
            if any(cell.value == "S. No." for cell in row)
        )
        assert header_values == [
            "S. No.",
            "Description of Goods / Materials",
            "Model / Part No.",
            "HSN/SAC",
            "Qty",
            "Unit",
            "Remarks",
        ]
        assert not any(
            phrase in values.lower()
            for phrase in ("vendor rate", "purchase rate", "taxable amount", "grand total")
        )
    assert (
        sheet.title
        == {
            "quotation": "Quotation",
            "purchase_order": "Purchase Order",
            "proforma_invoice": "Performa Invoice",
            "tax_invoice": "Tax Invoice",
            "challan": "Delivery Challan",
        }[kind]
    )
    assert sheet["A5"].fill.fgColor.rgb.endswith("5A2D0C")
    assert len(sheet._images) == 1


def test_stress_documents_and_indian_formatting():
    document = sample_document("tax_invoice", financial=True, stress=True)
    assert len(render_pdf(document)) > 10_000
    assert len(render_excel(document)) > 5_000
    assert indian_number("12345678.9") == "1,23,45,678.90"
    assert currency("125000") == "INR 1,25,000.00"
    assert safe_filename("QTN/000001 R1", "pdf") == "QTN-000001-R1.pdf"


def test_long_terms_are_limited_in_main_document_and_preserved_in_annexure():
    document = sample_document("tax_invoice", financial=True, stress=True)
    main_terms, annexure_terms = _presented_terms(document)
    assert len(main_terms) == 6
    assert main_terms[-1] == "Detailed Terms & Conditions: Refer Annexure-A."
    assert len(annexure_terms) == 160

    workbook = load_workbook(BytesIO(render_excel(document)), data_only=False)
    values = [
        str(cell.value)
        for row in workbook.active.iter_rows()
        for cell in row
        if cell.value is not None
    ]
    assert "ANNEXURE-A - DETAILED TERMS & CONDITIONS" in values
    signature = (
        "FOR WORLD COMMUNICATION\n\n\n____________________________\nAuthorised Signatory"
    )
    assert values.count(signature) == 2


@pytest.mark.parametrize(
    "kind", ["quotation", "proforma_invoice", "purchase_order", "tax_invoice", "challan"]
)
def test_normal_and_stress_documents_share_presentation_metadata(kind):
    normal = sample_document(kind, financial=kind != "challan")
    stress = sample_document(kind, financial=kind != "challan", stress=True)
    assert stress.identity == normal.identity
    assert stress.parties == normal.parties
    assert stress.continuation_reference == normal.continuation_reference
    assert render_pdf(normal).startswith(b"%PDF-")
    assert render_pdf(stress).startswith(b"%PDF-")


@pytest.mark.parametrize("kind", ["quotation", "proforma_invoice"])
def test_stress_documents_keep_final_page_one_metadata(kind):
    document = sample_document(kind, financial=True, stress=True)
    labels = {label.upper() for label, _ in document.identity}
    party_labels = {label.upper() for label, _ in document.parties}
    assert {"WORLD COMMUNICATION DOCUMENT NUMBER", "DATE"}.issubset(labels)
    assert {"CUSTOMER", "BILL TO", "SHIP TO"}.issubset(party_labels)
    if kind == "quotation":
        assert {"ENQUIRY / TENDER", "VALIDITY"}.issubset(labels)
    else:
        assert {"PROJECT NAME / LOA NUMBER", "RAILWAY DIVISION"}.issubset(labels)
        assert "RAILWAY AUTHORITY" in party_labels


def test_continuation_header_preserves_full_document_and_loa_references():
    document = sample_document("proforma_invoice", financial=True, stress=True)
    document.identifier = "PI-2026-VERY-LONG-BUT-AUTHORITATIVE-000001"
    document.continuation_reference = "LOA: ER/LOA/2026/004-LONG-AUTHORITATIVE-REFERENCE"
    document.identity = [
        (
            "PROJECT NAME / LOA NUMBER",
            "Extremely Long Railway Project Name That Must Remain Fully Visible On Page One / "
            "ER/LOA/2026/004-LONG-AUTHORITATIVE-REFERENCE",
        ),
        *document.identity,
    ]
    left, center, right = _continuation_parts(document)
    assert left == "WORLD COMMUNICATION"
    assert document.identifier in center
    assert right == document.continuation_reference
    assert "..." not in center + right
    assert "Extremely Long Railway Project Name" in document.identity[0][1]


def test_purchase_order_adapter_uses_vendor_and_preserved_oem_model_snapshots():
    line = SimpleNamespace(
        line_number=1,
        description="Managed switch",
        hsn_code="85176290",
        unit_snapshot="Nos",
        ordered_quantity="2",
        unit_rate="1000",
        discount_percent="0",
        taxable_amount="2000",
        cgst_percent="0",
        sgst_percent="0",
        igst_percent="18",
        line_total="2360",
        remarks=None,
        model_snapshot="C9200L-24P",
        oem_snapshot="Cisco",
    )
    record = SimpleNamespace(
        po_number="PO-000001",
        po_date="2026-08-28",
        project_id=uuid4(),
        loa_id=uuid4(),
        procurement_requirement_id=uuid4(),
        delivery_date="2026-09-30",
        status="ISSUED",
        organization_snapshot={"legal_name": "World Communication"},
        vendor_snapshot={
            "legal_name": "XYZ Distributor Pvt Ltd",
            "trade_name": "XYZ Distribution",
            "pan": "AAAAA0000A",
            "email": "sales@vendor.example",
        },
        billing_address_snapshot={"address_line_1": "Registered Office, Katihar"},
        shipping_address_snapshot={"address_line_1": "Railway Stores Depot, Howrah"},
        lines=[line],
        subtotal="2000",
        discount_amount="0",
        taxable_amount="2000",
        cgst_amount="0",
        sgst_amount="0",
        igst_amount="360",
        round_off="0",
        grand_total="2360",
        payment_terms_snapshot=None,
        terms_override_text=None,
        terms_snapshot=None,
        special_instructions=None,
    )
    document = purchase_order(record)
    assert [label for label, _ in document.parties] == [
        "VENDOR / SUPPLIER",
        "BUYER / BILL TO",
        "SHIP TO",
    ]
    assert "Customer" not in " ".join(label for label, _ in document.parties)
    assert "XYZ Distributor Pvt Ltd" in document.parties[0][1]
    assert document.lines[0].oem == "Cisco"
    assert document.lines[0].model == "C9200L-24P"


def test_long_terms_are_chunked_with_explicit_continuation_pages():
    document = sample_document("tax_invoice", financial=True, stress=True)
    assert len(_term_items(document.terms)) == 160
    flowables = _terms_flowables(
        document.terms,
        ParagraphStyle("Heading", textColor="white"),
        ParagraphStyle("Body"),
        700,
    )
    assert isinstance(flowables[0], PageBreak)
    assert sum(isinstance(item, PageBreak) for item in flowables) > 1
    headings = [
        item._content[0]._cellvalues[0][0].text
        for item in flowables
        if isinstance(item, KeepTogether) and item._content and isinstance(item._content[0], Table)
    ]
    assert headings[0] == "Terms & Conditions"
    assert all(value == "TERMS & CONDITIONS - CONTINUED" for value in headings[1:])


@pytest.mark.asyncio
async def test_export_endpoint_authentication_filename_and_content_type(
    client: AsyncClient, monkeypatch
):
    document_id = uuid4()
    assert (await client.get(f"/api/v1/documents/quotation/{document_id}/pdf")).status_code == 401
    token = await login(client, "admin@example.com", "admin-user-password")

    async def fake_export(self, document_type, record_id, output_format):
        assert document_type == "quotation" and record_id == document_id
        return b"%PDF-test", "application/pdf", "QTN-000001-R0.pdf"

    monkeypatch.setattr(DocumentExportService, "export", fake_export)
    response = await client.get(
        f"/api/v1/documents/quotation/{document_id}/pdf", headers=auth(token)
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.headers["content-disposition"] == 'attachment; filename="QTN-000001-R0.pdf"'
